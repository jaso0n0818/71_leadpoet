"""
Discovery UI — Filter leads, discover roles, collect unique companies.
Run: python ui.py
Open: http://localhost:5002
"""

import json
import os
import sys
import logging

from flask import Flask, request, jsonify, render_template_string

# Add project root to path so target_fit_model imports work
parent = os.path.dirname(__file__)
sys.path.insert(0, os.path.dirname(parent))
import importlib
if "target_fit_model" not in sys.modules:
    spec = importlib.util.spec_from_file_location(
        "target_fit_model",
        os.path.join(parent, "__init__.py"),
        submodule_search_locations=[parent],
    )
    mod = importlib.util.module_from_spec(spec)
    sys.modules["target_fit_model"] = mod
    spec.loader.exec_module(mod)

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(name)s %(levelname)s %(message)s")

from target_fit_model.query import fetch_unique_roles, fetch_leads_for_roles
from target_fit_model.scoring import compute_fit_score
from target_fit_model.role_discovery import score_roles_via_llm, get_ranked_roles
from target_fit_model.icp_extract import parse_icp as extract_icp_filters
from target_fit_model.config import COMPANY_MULTIPLIER, EMPLOYEE_RANGE_ORDER

app = Flask(__name__)
logger = logging.getLogger(__name__)

# ── Load geo data for dropdowns ──
_geo_path = os.path.join(parent, "geo_lookup_fast.json")
with open(_geo_path) as f:
    _GEO = json.load(f)

# All countries from geo_lookup_fast.json + "Any"
_SUPPORTED_COUNTRIES = ["Any"] + sorted([c.title() for c in _GEO.get("countries", [])])
_US_STATES = sorted([s.title() for s in _GEO.get("us_states", {}).keys()])
_CITIES_BY_COUNTRY = {}
for country, cities in _GEO.get("cities", {}).items():
    _CITIES_BY_COUNTRY[country.title()] = sorted([c.title() for c in cities])
_CITIES_BY_STATE = {}
for state, cities in _GEO.get("us_states", {}).items():
    _CITIES_BY_STATE[state.title()] = sorted([c.title() for c in cities])

# Employee count options — match target fit DB format exactly
_SIZES = ["2-10", "11-50", "51-200", "201-500", "501-1,000", "1,001-5,000", "5,001-10,000", "10,001+"]
_EMPLOYEE_VARIANTS = {
    "2-10": ["2-10", "2-10 employees"],
    "11-50": ["11-50", "11-50 employees"],
    "51-200": ["51-200", "51-200 employees"],
    "201-500": ["201-500", "201-500 employees"],
    "501-1,000": ["501-1,000", "501-1,000 employees"],
    "1,001-5,000": ["1,001-5,000", "1,001-5,000 employees"],
    "5,001-10,000": ["5,001-10,000", "5,001-10,000 employees"],
    "10,001+": ["10,001+", "10,001+ employees"],
}

# Load industries from taxonomy (same as target fit model)
sys.path.insert(0, parent)
from industry_taxonomy import INDUSTRY_TAXONOMY

_INDUSTRIES = sorted({ind for v in INDUSTRY_TAXONOMY.values() for ind in v["industries"]})

# Industry → sub-industry mapping from taxonomy
_INDUSTRY_SUBS = {}
for sub, data in INDUSTRY_TAXONOMY.items():
    for ind in data["industries"]:
        if ind not in _INDUSTRY_SUBS:
            _INDUSTRY_SUBS[ind] = []
        _INDUSTRY_SUBS[ind].append(sub)
for ind in _INDUSTRY_SUBS:
    _INDUSTRY_SUBS[ind] = sorted(set(_INDUSTRY_SUBS[ind]))

HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Discovery Pipeline</title>
<style>
  * { margin: 0; padding: 0; box-sizing: border-box; }
  body {
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    background: #0b1121; color: #e2e8f0; min-height: 100vh;
  }
  .container { max-width: 1200px; margin: 0 auto; padding: 30px 20px; }
  h1 { font-size: 24px; font-weight: 700; color: #fff; margin-bottom: 4px; }
  .subtitle { color: #4a5c73; font-size: 12px; margin-bottom: 24px; letter-spacing: 0.3px; }

  .form-grid {
    display: grid; grid-template-columns: 1fr 1fr; gap: 20px 16px; margin-bottom: 20px;
  }
  .form-group { display: flex; flex-direction: column; }
  .form-group.full-width { grid-column: 1 / -1; }
  .form-section-divider {
    grid-column: 1 / -1; height: 1px; background: #162040; margin: 4px 0;
  }
  label {
    font-size: 12px; font-weight: 600; color: #7a8ba0;
    text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 2px;
  }
  input, textarea {
    background: #0f1a2e; border: 1px solid #1c2d4a; border-radius: 8px;
    color: #e2e8f0; padding: 10px 14px; font-size: 14px;
    font-family: inherit; outline: none; transition: border-color 0.2s;
  }
  input:focus, textarea:focus { border-color: #3b7dd8; }
  textarea { resize: vertical; min-height: 80px; }

  /* Custom multi-select dropdown */
  .ms-wrap { position: relative; }
  .ms-input-area {
    background: #0f1a2e; border: 1px solid #1c2d4a; border-radius: 8px;
    min-height: 42px; display: flex; flex-wrap: wrap; align-items: center;
    gap: 5px; padding: 6px 10px; cursor: text; transition: border-color 0.2s;
  }
  .ms-input-area:hover { border-color: #263b5c; }
  .ms-input-area.focused { border-color: #3b7dd8; }
  .ms-pill {
    background: #162a4a; border: 1px solid #2a4a70; border-radius: 5px;
    padding: 3px 8px; font-size: 12px; color: #a8c0dc;
    display: flex; align-items: center; gap: 4px; white-space: nowrap;
    animation: pillIn 0.15s ease-out;
  }
  @keyframes pillIn { from { opacity: 0; transform: scale(0.85); } to { opacity: 1; transform: scale(1); } }
  .ms-pill .pill-x {
    cursor: pointer; color: #5a7a9a; font-size: 14px; line-height: 1;
    margin-left: 2px; border-radius: 3px; padding: 0 2px;
  }
  .ms-pill .pill-x:hover { color: #f55; background: rgba(255,80,80,0.1); }
  .ms-inline-input {
    flex: 1; min-width: 60px; background: transparent; border: none;
    color: #e2e8f0; font-size: 13px; outline: none; padding: 4px 2px;
    font-family: inherit;
  }
  .ms-inline-input::placeholder { color: #4a5c73; }
  .ms-dropdown {
    display: none; position: absolute; top: 100%; left: 0; right: 0;
    background: #0f1a2e; border: 1px solid #1c2d4a; border-radius: 8px;
    margin-top: 4px; max-height: 220px; overflow-y: auto; z-index: 100;
    box-shadow: 0 8px 24px rgba(0,0,0,0.5);
  }
  .ms-dropdown.open { display: block; }
  .ms-option {
    padding: 8px 12px; cursor: pointer; font-size: 13px;
    display: flex; align-items: center; gap: 8px; transition: background 0.1s;
  }
  .ms-option:hover { background: #152035; }
  .ms-option.selected { background: #162a4a; color: #5a9fd4; }
  .ms-option .check { width: 16px; color: #5a9fd4; font-weight: 700; font-size: 12px; }
  .ms-empty { padding: 10px 14px; color: #4a5c73; font-size: 13px; }

  .chip-container { display: flex; flex-wrap: wrap; gap: 6px; margin-top: 8px; min-height: 24px; }
  .chip {
    background: #162a4a; border: 1px solid #2a4a70; border-radius: 5px;
    padding: 3px 8px; font-size: 12px; color: #a8c0dc;
    display: flex; align-items: center; gap: 4px;
  }
  .chip.auto { border-color: #1a5a3a; background: #0f2a1a; }
  .chip .remove { cursor: pointer; color: #5a7a9a; font-size: 13px; line-height: 1; }
  .chip .remove:hover { color: #f55; }


  .queue-panel {
    margin-top: 16px; background: #0f1a2e; border: 1px solid #1c2d4a; border-radius: 10px;
    padding: 14px 18px;
  }
  .queue-header { display: flex; align-items: center; justify-content: space-between; margin-bottom: 10px; }
  .queue-title { font-size: 13px; font-weight: 600; color: #7a8ba0; text-transform: uppercase; letter-spacing: 0.5px; }
  .queue-count { font-size: 12px; color: #5a9fd4; font-weight: 600; }
  .queue-list { display: flex; flex-direction: column; gap: 6px; }
  .queue-item {
    display: flex; align-items: center; gap: 8px; padding: 8px 12px;
    background: #0b1121; border: 1px solid #1c2d4a; border-radius: 6px; font-size: 12px;
  }
  .queue-item.active { border-color: #3b7dd8; background: #111d35; }
  .queue-item.done { opacity: 0.5; }
  .queue-label { flex: 1; color: #c5d0dc; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
  .queue-meta { color: #5a7a9a; font-size: 11px; white-space: nowrap; }
  .queue-remove {
    color: #5a7a9a; cursor: pointer; font-size: 15px; line-height: 1; padding: 0 4px;
  }
  .queue-remove:hover { color: #f55; }
  .queue-actions { display: flex; gap: 8px; margin-top: 10px; }

  .btn {
    background: #3b7dd8; color: #fff; border: none; border-radius: 8px;
    padding: 12px 28px; font-size: 14px; font-weight: 600;
    cursor: pointer; transition: background 0.2s; margin-top: 8px;
  }
  .btn:hover { background: #2d6bc4; }
  .btn:disabled { background: #1c2d4a; color: #4a5c73; cursor: not-allowed; }
  .btn-row { display: flex; gap: 12px; align-items: center; }

  .phase {
    margin-top: 28px; background: #0f1a2e;
    border: 1px solid #1c2d4a; border-radius: 10px; overflow: hidden;
  }
  .phase-header {
    display: flex; align-items: center; gap: 10px;
    padding: 14px 18px; border-bottom: 1px solid #1c2d4a; background: #132038;
  }
  .phase-num {
    background: #3b7dd8; color: #fff; font-size: 11px;
    font-weight: 700; padding: 3px 8px; border-radius: 4px;
  }
  .phase-num.green { background: #1a5a3a; }
  .phase-title { font-size: 14px; font-weight: 600; color: #fff; }
  .phase-count { margin-left: auto; font-size: 13px; color: #5a9fd4; font-weight: 600; }
  .phase-body { padding: 16px 18px; max-height: 500px; overflow: auto; }

  .role-list { display: flex; flex-wrap: wrap; gap: 8px; }
  .role-chip {
    background: #162040; border: 1px solid #1c2d4a; border-radius: 6px;
    padding: 6px 12px; font-size: 13px; color: #c5d0dc;
  }
  .role-chip .rank { color: #5a9fd4; font-weight: 700; margin-right: 4px; }

  .companies-table { width: 100%; border-collapse: collapse; font-size: 13px; }
  .companies-table th {
    text-align: left; padding: 8px 10px; font-size: 11px; font-weight: 600;
    color: #5a6b80; text-transform: uppercase; letter-spacing: 0.5px;
    border-bottom: 1px solid #1c2d4a;
  }
  .companies-table td { padding: 10px; border-bottom: 1px solid #162040; vertical-align: top; }
  .companies-table tr:hover td { background: #121d33; }
  .company-name { font-weight: 600; color: #fff; }
  .fit-score { font-weight: 700; color: #5a9fd4; }
  .company-link { color: #5a9fd4; text-decoration: none; font-size: 12px; }
  .company-link:hover { text-decoration: underline; }
  .leads-count { background: #162040; border-radius: 4px; padding: 2px 8px; font-size: 12px; color: #7a8ba0; }
  .desc-cell {
    color: #6b7d93; font-size: 12px;
  }

  .spinner {
    display: inline-block; width: 16px; height: 16px;
    border: 2px solid #1c2d4a; border-top-color: #3b7dd8;
    border-radius: 50%; animation: spin 0.6s linear infinite;
    margin-right: 8px; vertical-align: middle;
  }
  @keyframes spin { to { transform: rotate(360deg); } }
  .status-msg { font-size: 13px; color: #7a8ba0; margin-left: 8px; }
  .hidden { display: none; }
  .help-text { font-size: 11px; color: #3d5069; margin-top: 2px; margin-bottom: 4px; line-height: 1.4; }

  .error-banner {
    margin-top: 12px; padding: 12px 16px; border-radius: 8px;
    background: #2a1215; border: 1px solid #5c2b2e; color: #e8a0a4;
    font-size: 13px; line-height: 1.5; display: flex; align-items: flex-start; gap: 10px;
  }
  .error-banner .error-icon { font-size: 16px; flex-shrink: 0; margin-top: 1px; }
  .error-banner .error-body { flex: 1; }
  .error-banner .error-title { font-weight: 600; color: #f0b0b3; margin-bottom: 2px; }
  .error-banner .error-dismiss {
    background: none; border: none; color: #8a5558; font-size: 16px;
    cursor: pointer; padding: 0 4px; line-height: 1; flex-shrink: 0;
  }
  .error-banner .error-dismiss:hover { color: #e8a0a4; }

  .warn-banner {
    margin-top: 12px; padding: 12px 16px; border-radius: 8px;
    background: #2a2215; border: 1px solid #5c4b2b; color: #e8d0a0;
    font-size: 13px; line-height: 1.5; display: flex; align-items: flex-start; gap: 10px;
  }
  .warn-banner .error-icon { font-size: 16px; flex-shrink: 0; margin-top: 1px; }
  .warn-banner .error-body { flex: 1; }
  .warn-banner .error-title { font-weight: 600; color: #f0d8a0; margin-bottom: 2px; }
  .warn-banner .error-dismiss {
    background: none; border: none; color: #8a7558; font-size: 16px;
    cursor: pointer; padding: 0 4px; line-height: 1; flex-shrink: 0;
  }
  .warn-banner .error-dismiss:hover { color: #e8d0a0; }

  .error-banner.hidden, .warn-banner.hidden { display: none; }

  .modal-overlay {
    display: none; position: fixed; top: 0; left: 0; right: 0; bottom: 0;
    background: rgba(5,10,20,0.8); z-index: 1000;
    align-items: center; justify-content: center;
  }
  .modal-overlay.open { display: flex; }
  .modal {
    background: #0f1a2e; border: 1px solid #1c2d4a; border-radius: 12px;
    width: 600px; max-width: 90vw; max-height: 80vh; overflow: hidden;
    box-shadow: 0 16px 48px rgba(0,0,0,0.6);
  }
  .modal-header {
    display: flex; align-items: center; justify-content: space-between;
    padding: 16px 20px; border-bottom: 1px solid #1c2d4a;
  }
  .modal-header h2 { font-size: 16px; color: #fff; margin: 0; }
  .modal-close {
    background: none; border: none; color: #5a6b80; font-size: 20px;
    cursor: pointer; padding: 4px 8px; line-height: 1;
  }
  .modal-close:hover { color: #fff; }
  .modal-body { padding: 20px; }
  .modal-body textarea {
    width: 100%; min-height: 200px; background: #121d33; border: 1px solid #1c2d4a;
    border-radius: 8px; color: #e2e8f0; padding: 14px; font-size: 13px;
    font-family: inherit; outline: none; resize: vertical;
  }
  .modal-body textarea:focus { border-color: #3b7dd8; }
  .modal-body p { font-size: 12px; color: #5a6b80; margin: 10px 0 0; }
  .modal-footer {
    display: flex; justify-content: flex-end; gap: 10px;
    padding: 14px 20px; border-top: 1px solid #1c2d4a;
  }
  .btn-secondary {
    background: #1c2d4a; color: #c5d0dc; border: none; border-radius: 8px;
    padding: 10px 20px; font-size: 13px; font-weight: 600; cursor: pointer;
  }
  .btn-secondary:hover { background: #263b5c; }
  .btn-paste {
    background: #1c2d4a; color: #c5d0dc; border: 1px solid #263b5c; border-radius: 8px;
    padding: 8px 16px; font-size: 13px; font-weight: 600; cursor: pointer;
    margin-top: 0; transition: all 0.2s;
  }
  .btn-paste:hover { background: #263b5c; border-color: #3b7dd8; color: #fff; }
</style>
</head>
<body>
<div class="container">
  <div style="display:flex;align-items:center;justify-content:space-between;">
    <h1>Discovery Pipeline</h1>
    <div style="display:flex;gap:10px;">
      <button type="button" class="btn-paste" onclick="openPasteModal()">Paste Brief</button>
      <button type="button" class="btn-paste" onclick="clearAllFields()" style="border-color:#3d5069;">Clear All</button>
    </div>
  </div>
  <p class="subtitle">Description &rarr; Industry extraction &rarr; DB query &rarr; Role ranking &rarr; Unique companies</p>

  <!-- Paste Brief Modal -->
  <div class="modal-overlay" id="pasteModal">
    <div class="modal">
      <div class="modal-header">
        <h2>Paste Brief</h2>
        <button class="modal-close" onclick="closePasteModal()">&times;</button>
      </div>
      <div class="modal-body">
        <textarea id="pasteBriefText" placeholder="Paste your brief here, e.g.:

Buyer Profile:
Head of Design, Head of Marketing at crypto companies.

Product/Service:
We help crypto companies with branding and design.

Intent Signals:
Raised funding in the last 30 days

Countries: United States
Company Sizes: 11-50, 51-200"></textarea>
        <p>Paste a brief and we'll auto-fill Buyer Profile, Product/Service, Intent Signals, Countries, and Company Sizes.</p>
      </div>
      <div class="modal-footer">
        <button class="btn-secondary" onclick="closePasteModal()">Cancel</button>
        <button class="btn" style="margin:0;" onclick="parsePastedBrief()">Fill Form</button>
      </div>
    </div>
  </div>

  <form id="discoveryForm">
    <div class="form-grid">

      <div class="form-group full-width">
        <label>Client Name</label>
        <p class="help-text">Used as the filename when results are saved. Leave blank to use buyer profile instead.</p>
        <input type="text" id="client_name" placeholder="e.g. Acme Corp, Q2 Campaign">
      </div>

      <div class="form-group full-width">
        <label>Buyer Profile</label>
        <p class="help-text">Describe the target companies and contacts you want to reach</p>
        <textarea id="request_description" placeholder="Head of Design, Head of Marketing at small and mid-size crypto companies."></textarea>
      </div>

      <div class="form-group full-width">
        <label>Product/Service</label>
        <p class="help-text">What you sell &mdash; helps distinguish your offering from the target company's industry</p>
        <textarea id="product_description" placeholder="We help crypto companies with branding, web and product design."></textarea>
      </div>

      <div class="form-group full-width">
        <label>Intent Signals</label>
        <p class="help-text">Behavioral signals that suggest a company may need your product right now</p>
        <textarea id="intent_signals" placeholder="Raised funding in the last 30 days, hiring for designers or heads of marketing/brand."></textarea>
      </div>

      <div class="form-group full-width">
        <label>Lead Source</label>
        <div style="display:flex;gap:20px;">
          <label style="font-weight:400;cursor:pointer;"><input type="radio" name="lead_source" value="fundable" checked onchange="toggleLeadSource()"> Fundable</label>
        </div>
        <input type="hidden" name="lead_source" value="fundable">
      </div>

      <div id="fundableFields" style="display:none;">
        <div class="form-group" style="display:inline-block;width:48%;margin-right:4%;">
          <label>Funding Recency (days)</label>
          <p class="help-text">Only companies funded in last X days (leave blank for any)</p>
          <input type="number" id="funding_days" placeholder="e.g. 90" min="1" max="730">
        </div>
        <div class="form-group" style="display:inline-block;width:48%;">
          <label>Financing Types</label>
          <p class="help-text">e.g. SEED, SERIES_A, SERIES_B, EQUITY</p>
          <input type="text" id="financing_types" placeholder="e.g. SEED,SERIES_A">
        </div>
      </div>

      <div class="form-section-divider"></div>

      <div class="form-group">
        <label>Industry <span style="font-weight:400;text-transform:none;color:#4a5c73">(auto or manual)</span></label>
        <p class="help-text">Auto-extracted via Claude from the fields above, or select manually to override</p>
        <div class="ms-wrap" id="industryWrap"></div>
      </div>

      <div class="form-group">
        <label>Sub-Industry <span style="font-weight:400;text-transform:none;color:#4a5c73">(auto or manual)</span></label>
        <p class="help-text">Auto-populated from selected industries, or adjust manually</p>
        <div class="ms-wrap" id="subIndustryWrap"></div>
      </div>

      <div class="form-group full-width">
        <label>Role <span style="font-weight:400;text-transform:none;color:#4a5c73">(auto or manual)</span></label>
        <p class="help-text">Auto-discovered from buyer profile, or add specific roles to prioritize in ranking</p>
        <div style="display:flex;gap:8px;align-items:center;">
          <input type="text" id="roleInput" placeholder="Type a role and press Enter" style="flex:1;" onkeydown="if(event.key==='Enter'){event.preventDefault();addRole();}">
          <button type="button" class="btn" style="margin:0;padding:8px 16px;font-size:13px;" onclick="addRole()">Add</button>
        </div>
        <div class="chip-container" id="roleChips"></div>
      </div>

      <div class="form-section-divider"></div>

      <div class="form-group">
        <label>Company Sizes</label>
        <div class="ms-wrap" id="sizesWrap"></div>
      </div>

      <div class="form-group">
        <label>Countries</label>
        <div class="ms-wrap" id="countriesWrap"></div>
      </div>

      <div class="form-group">
        <label>States</label>
        <div class="ms-wrap" id="statesWrap"></div>
      </div>

      <div class="form-group">
        <label>Cities</label>
        <div class="ms-wrap" id="citiesWrap"></div>
      </div>

      <div class="form-section-divider"></div>

      <div class="form-group" style="grid-column: 1 / -1;">
        <label>Exclude Companies</label>
        <p class="help-text">Paste company names (one per line) to exclude from results</p>
        <div class="ms-input-area" id="excludeArea">
          <input class="ms-inline-input" id="excludeInput" placeholder="Paste company names here, one per line...">
        </div>
      </div>

      <div class="form-section-divider"></div>

      <div class="form-group">
        <label>Max Leads</label>
        <input type="number" id="max_leads" value="5" min="1">
      </div>
    </div>

    <div class="btn-row">
      <button type="submit" class="btn" id="runBtn">Run Discovery</button>
      <button type="button" class="btn" id="addQueueBtn" onclick="addToQueue()" style="background:#2a4a6a;font-size:12px;padding:10px 18px;">Add to Queue</button>
      <button type="button" class="btn hidden" id="cancelBtn" style="background:#6b2a2a;">Cancel</button>
      <label style="font-size:12px;color:#7a8ba0;margin-left:12px;display:flex;align-items:center;gap:4px;cursor:pointer;"><input type="checkbox" id="fullRunCheck" style="margin:0;">Full Run</label>
      <span id="loading" class="hidden"><span class="spinner"></span><span class="status-msg" id="statusMsg">Extracting industries...</span></span>
    </div>
    <div id="errorBanner" class="error-banner hidden"></div>
  </form>

  <div id="queuePanel" class="queue-panel hidden">
    <div class="queue-header">
      <span class="queue-title">Queue <span id="queueCount" class="queue-count"></span></span>
    </div>
    <div class="queue-list" id="queueList"></div>
    <div class="queue-actions">
      <button type="button" class="btn" id="runQueueBtn" onclick="processQueue()" style="background:#2a6a2a;font-size:12px;padding:8px 18px;">Run Queue</button>
      <button type="button" class="btn hidden" id="stopQueueBtn" onclick="stopQueue()" style="background:#6b2a2a;font-size:12px;padding:8px 18px;">Stop Queue</button>
      <button type="button" class="btn" id="clearQueueBtn" onclick="clearQueue()" style="background:#555;font-size:12px;padding:8px 14px;">Clear</button>
    </div>
  </div>

  <!-- Phase 0: Extracted Industries -->
  <div id="phase0" class="phase hidden">
    <div class="phase-header">
      <span class="phase-num green">0</span>
      <span class="phase-title">Extracted Industries (LLM)</span>
      <span class="phase-count" id="phase0Count"></span>
    </div>
    <div class="phase-body" id="phase0Body"></div>
  </div>

  <!-- Phase 1: DB Results -->
  <div id="phase1" class="phase hidden">
    <div class="phase-header">
      <span class="phase-num">1</span>
      <span class="phase-title">DB Query Results</span>
      <span class="phase-count" id="phase1Count"></span>
    </div>
    <div class="phase-body"><p id="phase1Info" style="font-size:13px;color:#6b7d93;"></p></div>
  </div>

  <!-- Phase 2: Ranked Roles -->
  <div id="phase2" class="phase hidden">
    <div class="phase-header">
      <span class="phase-num">2</span>
      <span class="phase-title">Ranked Roles (LLM)</span>
      <span class="phase-count" id="phase2Count"></span>
    </div>
    <div class="phase-body"><div class="role-list" id="roleList"></div></div>
  </div>

  <!-- Phase 3: Unique Companies -->
  <div id="phase3" class="phase hidden">
    <div class="phase-header">
      <span class="phase-num">3</span>
      <span class="phase-title">Unique Companies</span>
      <span class="phase-count" id="phase3Count"></span>
    </div>
    <div class="phase-body" style="max-height:700px;">
      <table class="companies-table">
        <thead>
          <tr><th>#</th><th>Company</th><th>Location</th><th>Website</th><th>LinkedIn</th><th>Size</th><th>Industry</th><th>Description</th><th>Roles</th><th>Fit</th><th>Leads</th></tr>
        </thead>
        <tbody id="companiesBody"></tbody>
      </table>
    </div>
    <div style="padding:16px 18px;border-top:1px solid #1c2d4a;">
      <div class="btn-row">
        <button type="button" class="btn" onclick="downloadTableCSV('phase3')" style="background:#555;font-size:12px;padding:8px 14px;">Download Companies CSV</button>
        <button type="button" class="btn" onclick="downloadAllLeadsExcel()" style="background:#555;font-size:12px;padding:8px 14px;">Download All Leads Excel</button>
        <label style="font-size:12px;color:#7a8ba0;margin-right:12px;"><input type="checkbox" id="skipStage4" style="margin-right:4px;">Skip Stage 4</label>
        <button type="button" class="btn" id="processBtn" onclick="runProcessCompanies()" style="background:#2a6a2a;">Verify + Intent Check</button>
        <button type="button" class="btn hidden" id="processCancelBtn" style="background:#6b2a2a;font-size:12px;padding:8px 14px;">Cancel</button>
        <span id="processLoading" class="hidden"><span class="spinner"></span><span class="status-msg" id="processStatusMsg">Processing...</span></span>
      </div>
    </div>
  </div>

  <!-- Funding Check Results -->
  <div id="fundingPhase" class="phase hidden">
    <div class="phase-header">
      <span class="phase-num" style="background:#d4a017;">F</span>
      <span class="phase-title">Funding Check</span>
      <span class="phase-count" id="fundingCount"></span>
    </div>
    <div class="phase-body">
      <table class="companies-table">
        <thead><tr><th>#</th><th>Company</th><th>Status</th><th>Evidence</th></tr></thead>
        <tbody id="fundingBody"></tbody>
      </table>
    </div>
  </div>

  <!-- Phase 4: Intent Results -->
  <div id="phase4" class="phase hidden">
    <div class="phase-header">
      <span class="phase-num green">4</span>
      <span class="phase-title">Intent Check Results</span>
      <span class="phase-count" id="phase4Count"></span>
    </div>
    <div class="phase-body" style="max-height:700px;">
      <div style="margin-bottom:8px;">
        <button type="button" class="btn" onclick="downloadTableCSV('phase4')" style="background:#555;font-size:12px;padding:8px 14px;">Download CSV</button>
      </div>
      <table class="companies-table">
        <thead>
          <tr><th>#</th><th>Company</th><th>Location</th><th>Website</th><th>LinkedIn</th><th>Size</th><th>Industry</th><th>Sub-Industry</th><th>Description</th><th>Roles</th><th>Intent Details</th><th>Intent Score</th><th>Lead Score</th></tr>
        </thead>
        <tbody id="intentBody"></tbody>
      </table>
    </div>
  </div>

  <!-- Phase 5: Final Leads + Prospect Pool -->
  <div id="phase5" class="phase hidden">
    <div class="phase-header">
      <span class="phase-num" style="background:#c90;">5</span>
      <span class="phase-title">Final Leads</span>
      <span class="phase-count" id="phase5Count"></span>
    </div>
    <div class="phase-body" style="max-height:700px;">
      <div style="margin-bottom:12px;">
        <button type="button" class="btn" onclick="downloadExcel()" style="background:#d4950a;color:#fff;">Download Excel (Both Sheets)</button>
      </div>
      <h3 style="color:#5a9fd4;margin:16px 0 8px;">Leads (Intent Score &ge; 30%)</h3>
      <table class="companies-table">
        <thead>
          <tr><th>#</th><th>Name</th><th>Email</th><th>Role</th><th>Company</th><th>LinkedIn</th><th>Website</th><th>Company LinkedIn</th><th>Phone</th><th>Industry</th><th>Sub Industry</th><th>City</th><th>State</th><th>Country</th><th>HQ City</th><th>HQ State</th><th>HQ Country</th><th>Size</th><th>Description</th><th>Intent Details</th><th>Lead Score</th><th>Evidence</th></tr>
        </thead>
        <tbody id="leadsPoolBody"></tbody>
      </table>
      <h3 style="color:#7a8ba0;margin:24px 0 8px;cursor:pointer;" onclick="document.getElementById('prospectSection').classList.toggle('hidden')">Prospect Pool (Intent Score &lt; 30%) &#9662;</h3>
      <div id="prospectSection" class="hidden">
        <table class="companies-table">
          <thead>
            <tr><th>#</th><th>Name</th><th>Email</th><th>Role</th><th>Company</th><th>LinkedIn</th><th>Website</th><th>Company LinkedIn</th><th>Phone</th><th>Industry</th><th>Sub Industry</th><th>City</th><th>State</th><th>Country</th><th>HQ City</th><th>HQ State</th><th>HQ Country</th><th>Size</th><th>Description</th><th>Intent Details</th><th>Lead Score</th><th>Evidence</th></tr>
          </thead>
          <tbody id="prospectPoolBody"></tbody>
        </table>
      </div>
    </div>
  </div>
</div>

<script>
// ── Multi-Select Component ──
class MultiSelect {
  constructor(wrapId, options, placeholder, onChange) {
    this.wrap = document.getElementById(wrapId);
    this.allOptions = options;
    this.selected = [];
    this.placeholder = placeholder;
    this.onChange = onChange || (() => {});
    this.isOpen = false;
    this.searchVal = '';
    this.build();

    document.addEventListener('mousedown', (e) => {
      if (this.isOpen && !this.wrap.contains(e.target)) {
        this.close();
      }
    });
  }

  build() {
    this.wrap.innerHTML = '';

    this.inputArea = document.createElement('div');
    this.inputArea.className = 'ms-input-area';
    this.inputArea.addEventListener('mousedown', (e) => {
      if (e.target === this.inputArea) {
        e.preventDefault();
        this.isOpen ? this.close() : this.openDd();
      }
    });
    this.wrap.appendChild(this.inputArea);

    this.searchInput = document.createElement('input');
    this.searchInput.className = 'ms-inline-input';
    this.searchInput.placeholder = this.placeholder;
    this.searchInput.addEventListener('input', () => {
      this.searchVal = this.searchInput.value;
      this.updateOptions();
      if (!this.isOpen && this.searchVal) this.openDd();
    });
    this.searchInput.addEventListener('focus', () => {
      this.inputArea.classList.add('focused');
      if (!this.isOpen) this.openDd();
    });
    this.searchInput.addEventListener('blur', () => {
      this.inputArea.classList.remove('focused');
    });
    this.searchInput.addEventListener('keydown', (e) => {
      if (e.key === 'Backspace' && !this.searchInput.value && this.selected.length) {
        e.preventDefault();
        this.selected.pop();
        this.refresh();
        this.onChange(this.selected);
      }
    });
    this.searchInput.addEventListener('mousedown', (e) => e.stopPropagation());
    this.inputArea.appendChild(this.searchInput);

    this.dd = document.createElement('div');
    this.dd.className = 'ms-dropdown';
    this.optionsContainer = document.createElement('div');
    this.dd.appendChild(this.optionsContainer);
    this.wrap.appendChild(this.dd);

    this.renderPills();
    this.updateOptions();
  }

  renderPills() {
    const pills = this.inputArea.querySelectorAll('.ms-pill');
    pills.forEach(p => p.remove());

    this.selected.forEach((val, i) => {
      const pill = document.createElement('span');
      pill.className = 'ms-pill';
      pill.textContent = val;
      const x = document.createElement('span');
      x.className = 'pill-x';
      x.innerHTML = '&times;';
      x.addEventListener('mousedown', (e) => {
        e.preventDefault();
        e.stopPropagation();
        this.selected.splice(i, 1);
        this.refresh();
        this.onChange(this.selected);
      });
      pill.appendChild(x);
      this.inputArea.insertBefore(pill, this.searchInput);
    });

    this.searchInput.placeholder = this.selected.length ? '' : this.placeholder;
  }

  updateOptions() {
    this.optionsContainer.innerHTML = '';
    const lower = this.searchVal.toLowerCase();
    const filtered = this.allOptions.filter(o => o.toLowerCase().includes(lower));

    if (filtered.length === 0) {
      const empty = document.createElement('div');
      empty.className = 'ms-empty';
      empty.textContent = 'No matches';
      this.optionsContainer.appendChild(empty);
      return;
    }

    filtered.forEach(opt => {
      const div = document.createElement('div');
      const isSel = this.selected.includes(opt);
      div.className = 'ms-option' + (isSel ? ' selected' : '');
      div.innerHTML = '<span class="check">' + (isSel ? '&#10003;' : '') + '</span>' + opt;
      div.addEventListener('mousedown', (e) => {
        e.preventDefault();
        e.stopPropagation();
        if (this.selected.includes(opt)) {
          this.selected = this.selected.filter(s => s !== opt);
        } else {
          this.selected.push(opt);
        }
        this.refresh();
        this.onChange(this.selected);
        this.searchInput.value = '';
        this.searchVal = '';
        this.updateOptions();
        this.searchInput.focus();
      });
      this.optionsContainer.appendChild(div);
    });
  }

  refresh() {
    this.renderPills();
    this.updateOptions();
  }

  openDd() {
    this.isOpen = true;
    this.dd.classList.add('open');
    this.searchInput.value = '';
    this.searchVal = '';
    this.updateOptions();
    setTimeout(() => this.searchInput.focus(), 10);
  }

  close() {
    this.isOpen = false;
    this.dd.classList.remove('open');
  }

  setSelected(vals) {
    this.selected = [...vals];
    this.refresh();
  }

  setOptions(opts) {
    this.allOptions = opts;
    this.selected = this.selected.filter(s => opts.includes(s));
    this.refresh();
  }

  getSelected() { return [...this.selected]; }
}

// ── Data ──
const ALL_COUNTRIES = """ + json.dumps(_SUPPORTED_COUNTRIES) + """;
const ALL_STATES = """ + json.dumps(_US_STATES) + """;
const CITIES_BY_COUNTRY = """ + json.dumps(_CITIES_BY_COUNTRY) + """;
const CITIES_BY_STATE = """ + json.dumps(_CITIES_BY_STATE) + """;
const ALL_SIZES = """ + json.dumps(_SIZES) + """;
const ALL_INDUSTRIES = """ + json.dumps(_INDUSTRIES) + """;
const INDUSTRY_SUBS = """ + json.dumps(_INDUSTRY_SUBS) + """;

// ── Init Multi-Selects ──
const sizesMS = new MultiSelect('sizesWrap', ALL_SIZES, 'Select company sizes...');
const countriesMS = new MultiSelect('countriesWrap', ALL_COUNTRIES, 'Select countries...', onCountryChange);
countriesMS.setSelected(['Any']);
const statesMS = new MultiSelect('statesWrap', [], 'Select states...', onStateChange);
const citiesMS = new MultiSelect('citiesWrap', [], 'Select cities...');
const industryMS = new MultiSelect('industryWrap', ALL_INDUSTRIES, 'Select or auto-detect...', onIndustryChange);
const subIndustryMS = new MultiSelect('subIndustryWrap', [], 'Auto-expanded from industries...');

// ── Lead source toggle ──
function toggleLeadSource() {
  const source = document.querySelector('input[name="lead_source"]:checked')?.value || 'supabase';
  document.getElementById('fundableFields').style.display = source === 'fundable' ? 'block' : 'none';
}

// ── Role chip management ──
let selectedRoles = [];

function addRole() {
  const input = document.getElementById('roleInput');
  const val = input.value.trim();
  if (val && !selectedRoles.includes(val)) {
    selectedRoles.push(val);
    renderRoleChips();
  }
  input.value = '';
  input.focus();
}

function removeRole(i) {
  selectedRoles.splice(i, 1);
  renderRoleChips();
}

function renderRoleChips() {
  const container = document.getElementById('roleChips');
  container.innerHTML = '';
  selectedRoles.forEach((role, i) => {
    const chip = document.createElement('div');
    chip.className = 'chip';
    chip.textContent = role;
    const rm = document.createElement('span');
    rm.className = 'remove';
    rm.innerHTML = '&times;';
    rm.addEventListener('mousedown', (e) => {
      e.preventDefault();
      removeRole(i);
    });
    chip.appendChild(rm);
    container.appendChild(chip);
  });
}

// ── Exclude Companies ──
let excludedCompanies = [];
const excludeArea = document.getElementById('excludeArea');
const excludeInput = document.getElementById('excludeInput');

function addExcludeCompany(name) {
  const trimmed = name.trim();
  if (trimmed && !excludedCompanies.includes(trimmed)) {
    excludedCompanies.push(trimmed);
  }
}

function removeExcludeCompany(i) {
  excludedCompanies.splice(i, 1);
  renderExcludeChips();
  try { saveFormState(); } catch(e) {}
}

function renderExcludeChips() {
  excludeArea.querySelectorAll('.ms-pill').forEach(p => p.remove());
  excludedCompanies.forEach((name, i) => {
    const pill = document.createElement('span');
    pill.className = 'ms-pill';
    pill.textContent = name;
    const x = document.createElement('span');
    x.className = 'pill-x';
    x.innerHTML = '&times;';
    x.addEventListener('mousedown', (e) => {
      e.preventDefault();
      e.stopPropagation();
      removeExcludeCompany(i);
    });
    pill.appendChild(x);
    excludeArea.insertBefore(pill, excludeInput);
  });
  excludeInput.placeholder = excludedCompanies.length ? '' : 'Paste company names here, one per line...';
}

excludeArea.addEventListener('mousedown', (e) => {
  if (e.target === excludeArea) { e.preventDefault(); excludeInput.focus(); }
});

excludeInput.addEventListener('paste', (e) => {
  e.preventDefault();
  const text = (e.clipboardData || window.clipboardData).getData('text');
  const lines = text.split(/\\r?\\n/).map(l => l.trim()).filter(Boolean);
  if (lines.length > 0) {
    lines.forEach(l => addExcludeCompany(l));
    renderExcludeChips();
    excludeInput.value = '';
    try { saveFormState(); } catch(ex) {}
  }
});

excludeInput.addEventListener('keydown', (e) => {
  if (e.key === 'Enter') {
    e.preventDefault();
    const val = excludeInput.value.trim();
    if (val) {
      addExcludeCompany(val);
      renderExcludeChips();
      excludeInput.value = '';
      try { saveFormState(); } catch(ex) {}
    }
  }
  if (e.key === 'Backspace' && !excludeInput.value && excludedCompanies.length) {
    e.preventDefault();
    excludedCompanies.pop();
    renderExcludeChips();
    try { saveFormState(); } catch(ex) {}
  }
});

function onIndustryChange(selected) {
  const validSubs = new Set();
  selected.forEach(ind => {
    (INDUSTRY_SUBS[ind] || []).forEach(s => validSubs.add(s));
  });
  const newOptions = [...validSubs].sort();
  const currentlySelected = subIndustryMS.getSelected();
  const kept = currentlySelected.filter(s => validSubs.has(s));
  subIndustryMS.setOptions(newOptions);
  if (kept.length > 0) {
    subIndustryMS.setSelected(kept);
  } else {
    subIndustryMS.setSelected(newOptions);
  }
}

function onCountryChange(selected) {
  if (selected.length === 0) {
    countriesMS.setSelected(['Any']);
    return;
  }
  // "Any" means global — no state/city filtering
  if (selected.includes('Any')) {
    statesMS.setOptions([]);
    citiesMS.setOptions([]);
    return;
  }
  // States only available for US
  if (selected.includes('United States')) {
    statesMS.setOptions(ALL_STATES);
  } else {
    statesMS.setOptions([]);
  }
  updateCityOptions();
}

function onStateChange(selected) {
  updateCityOptions();
}

function updateCityOptions() {
  const cities = new Set();
  const selStates = statesMS.getSelected();
  const selCountries = countriesMS.getSelected();

  selStates.forEach(s => {
    (CITIES_BY_STATE[s] || []).forEach(c => cities.add(c));
  });
  if (selStates.length === 0) {
    selCountries.forEach(c => {
      (CITIES_BY_COUNTRY[c] || []).forEach(ct => cities.add(ct));
    });
  }
  citiesMS.setOptions([...cities].sort());
}

// ── Error/Warning Display ──
const errorBanner = document.getElementById('errorBanner');

function showError(title, detail) {
  errorBanner.className = 'error-banner';
  errorBanner.innerHTML =
    '<span class="error-icon">&#9888;</span>' +
    '<div class="error-body"><div class="error-title">' + title + '</div>' +
    (detail ? '<div>' + detail + '</div>' : '') + '</div>' +
    '<button class="error-dismiss" onclick="dismissError()">&times;</button>';
}

function showWarning(title, detail) {
  errorBanner.className = 'warn-banner';
  errorBanner.innerHTML =
    '<span class="error-icon">&#9888;</span>' +
    '<div class="error-body"><div class="error-title">' + title + '</div>' +
    (detail ? '<div>' + detail + '</div>' : '') + '</div>' +
    '<button class="error-dismiss" onclick="dismissError()">&times;</button>';
}

function dismissError() {
  errorBanner.className = 'error-banner hidden';
  errorBanner.innerHTML = '';
}

// (stale localStorage cleanup removed — persistence now active)

// ── LocalStorage Persistence ──
const STORAGE_KEY = 'discovery_form_v2';

function captureFormState() {
  return {
    sizes: sizesMS.getSelected(),
    countries: countriesMS.getSelected(),
    states: statesMS.getSelected(),
    cities: citiesMS.getSelected(),
    industries: industryMS.getSelected(),
    subIndustries: subIndustryMS.getSelected(),
    roles: selectedRoles.slice(),
    excludeCompanies: excludedCompanies.slice(),
    clientName: document.getElementById('client_name').value,
    buyer: document.getElementById('request_description').value,
    product: document.getElementById('product_description').value,
    intent: document.getElementById('intent_signals').value,
    maxLeads: document.getElementById('max_leads').value,
    leadSource: document.querySelector('input[name="lead_source"]:checked')?.value || 'supabase',
    fullRun: document.getElementById('fullRunCheck').checked,
  };
}

function saveFormState() {
  try { localStorage.setItem(STORAGE_KEY, JSON.stringify(captureFormState())); } catch(e) {}
}

function applyFormState(s) {
  try {
    document.getElementById('client_name').value = s.clientName || '';
    document.getElementById('request_description').value = s.buyer || '';
    document.getElementById('product_description').value = s.product || '';
    document.getElementById('intent_signals').value = s.intent || '';
    document.getElementById('max_leads').value = s.maxLeads || '5';

    countriesMS.setSelected(s.countries?.length ? s.countries : ['Any']);
    onCountryChange(s.countries?.length ? s.countries : ['Any']);

    if (s.states?.length) {
      statesMS.setSelected(s.states);
      onStateChange(s.states);
    } else { statesMS.setSelected([]); }

    if (s.cities?.length) {
      updateCityOptions();
      citiesMS.setSelected(s.cities);
    } else { citiesMS.setSelected([]); }

    sizesMS.setSelected(s.sizes || []);

    if (s.industries?.length) {
      industryMS.setSelected(s.industries);
      onIndustryChange(s.industries);
    } else {
      industryMS.setSelected([]);
      subIndustryMS.setOptions([]);
      subIndustryMS.setSelected([]);
    }
    if (s.subIndustries?.length) subIndustryMS.setSelected(s.subIndustries);

    selectedRoles = (s.roles || []).slice();
    renderRoleChips();

    excludedCompanies = (s.excludeCompanies || []).slice();
    renderExcludeChips();

    if (s.leadSource) {
      const radio = document.querySelector('input[name="lead_source"][value="' + s.leadSource + '"]');
      if (radio) { radio.checked = true; toggleLeadSource(); }
    }
    document.getElementById('fullRunCheck').checked = !!s.fullRun;
  } catch(e) { console.warn('Failed to apply form state', e); }
}

function restoreFormState() {
  try {
    const raw = localStorage.getItem(STORAGE_KEY);
    if (!raw) return;
    applyFormState(JSON.parse(raw));
  } catch(e) { console.warn('Failed to restore form state', e); }
}

restoreFormState();

document.querySelectorAll('input, textarea').forEach(el => {
  el.addEventListener('input', saveFormState);
});

const _origAddRole = addRole;
addRole = function() { _origAddRole(); saveFormState(); };
const _origRemoveRole = removeRole;
removeRole = function(i) { _origRemoveRole(i); saveFormState(); };

[sizesMS, countriesMS, statesMS, citiesMS, industryMS, subIndustryMS].forEach(ms => {
  const _origRefresh = ms.refresh.bind(ms);
  ms.refresh = function() { _origRefresh(); try { saveFormState(); } catch(e) {} };
});

// ── Queue System ──
const QUEUE_KEY = 'discovery_queue';
let requestQueue = [];
let queueRunning = false;
let queueCancelled = false;
let queueCompleted = 0;
let queueTotal = 0;

function saveQueue() {
  try { localStorage.setItem(QUEUE_KEY, JSON.stringify(requestQueue)); } catch(e) {}
}

function loadQueue() {
  try {
    const raw = localStorage.getItem(QUEUE_KEY);
    if (raw) requestQueue = JSON.parse(raw);
  } catch(e) { requestQueue = []; }
}

function addToQueue() {
  const state = captureFormState();
  const desc = (state.clientName || state.buyer || state.product || '').trim();
  state._label = desc ? desc.substring(0, 60) : 'Untitled request';
  state._id = Date.now() + '_' + Math.random().toString(36).substr(2, 5);
  requestQueue.push(state);
  saveQueue();
  renderQueuePanel();
}

function removeFromQueue(idx) {
  if (queueRunning && idx === 0) return;
  requestQueue.splice(idx, 1);
  saveQueue();
  renderQueuePanel();
}

function clearQueue() {
  if (queueRunning) return;
  requestQueue = [];
  saveQueue();
  renderQueuePanel();
}

function renderQueuePanel() {
  const panel = document.getElementById('queuePanel');
  if (requestQueue.length === 0 && !queueRunning) {
    panel.classList.add('hidden');
    return;
  }
  panel.classList.remove('hidden');

  const countEl = document.getElementById('queueCount');
  const listEl = document.getElementById('queueList');
  const runQueueBtn = document.getElementById('runQueueBtn');
  const stopQueueBtn = document.getElementById('stopQueueBtn');
  const clearQueueBtn = document.getElementById('clearQueueBtn');

  if (queueRunning) {
    const remaining = requestQueue.length;
    const total = queueCompleted + remaining;
    countEl.textContent = '— running ' + (queueCompleted + 1) + '/' + total;
    runQueueBtn.classList.add('hidden');
    clearQueueBtn.classList.add('hidden');
    stopQueueBtn.classList.remove('hidden');
  } else {
    countEl.textContent = '— ' + requestQueue.length + ' item' + (requestQueue.length === 1 ? '' : 's');
    runQueueBtn.classList.remove('hidden');
    clearQueueBtn.classList.remove('hidden');
    stopQueueBtn.classList.add('hidden');
  }

  listEl.innerHTML = '';
  requestQueue.forEach((item, i) => {
    const div = document.createElement('div');
    div.className = 'queue-item' + (queueRunning && i === 0 ? ' active' : '');
    const label = item._label || 'Untitled';
    const leads = item.maxLeads || '5';
    div.innerHTML = '<span class="queue-label">' + (queueRunning && i === 0 ? '▸ ' : '') + label + '</span>'
      + '<span class="queue-meta">' + leads + ' leads</span>'
      + (queueRunning && i === 0 ? '' : '<span class="queue-remove" onclick="removeFromQueue(' + i + ')">&times;</span>');
    listEl.appendChild(div);
  });
}

async function processQueue() {
  if (requestQueue.length === 0) return;
  queueRunning = true;
  queueCancelled = false;
  queueCompleted = 0;
  queueTotal = requestQueue.length;
  renderQueuePanel();

  while (requestQueue.length > 0 && !queueCancelled) {
    const item = requestQueue[0];
    renderQueuePanel();

    applyFormState(item);
    document.getElementById('fullRunCheck').checked = true;

    await new Promise(resolve => {
      document.addEventListener('runComplete', resolve, { once: true });
      form.dispatchEvent(new Event('submit', { bubbles: true, cancelable: true }));
    });

    if (queueCancelled) break;

    requestQueue.shift();
    queueCompleted++;
    saveQueue();
    renderQueuePanel();

    if (requestQueue.length > 0 && !queueCancelled) {
      await new Promise(r => setTimeout(r, 3000));
    }
  }

  queueRunning = false;
  queueCancelled = false;
  renderQueuePanel();
  if (requestQueue.length === 0) {
    showWarning('Queue complete', 'All ' + queueTotal + ' queued request' + (queueTotal === 1 ? '' : 's') + ' have been processed.');
  }
}

function stopQueue() {
  queueCancelled = true;
  if (discoverAbort) { discoverAbort.abort(); discoverAbort = null; }
  if (window._processAbort) { window._processAbort.abort(); window._processAbort = null; }
  const total = queueCompleted + requestQueue.length;
  showWarning('Queue stopped', 'Completed ' + queueCompleted + '/' + total + '. Remaining items preserved in queue.');
}

loadQueue();
renderQueuePanel();

// ── Form Submit ──
const form = document.getElementById('discoveryForm');
const loading = document.getElementById('loading');
const statusMsg = document.getElementById('statusMsg');
const runBtn = document.getElementById('runBtn');
const cancelBtn = document.getElementById('cancelBtn');
let discoverAbort = null;

cancelBtn.addEventListener('click', () => {
  if (discoverAbort) { discoverAbort.abort(); discoverAbort = null; }
  showWarning('Cancelled', 'Discovery run was cancelled.');
  statusMsg.textContent = 'Cancelled';
  runBtn.disabled = false;
  cancelBtn.classList.add('hidden');
  loading.classList.add('hidden');
});

form.addEventListener('submit', async (e) => {
  e.preventDefault();
  if (discoverAbort) { discoverAbort.abort(); }
  discoverAbort = new AbortController();
  runBtn.disabled = true;
  cancelBtn.classList.remove('hidden');
  loading.classList.remove('hidden');
  dismissError();
  statusMsg.textContent = 'Extracting industries...';

  ['phase0','phase1','phase2','phase3'].forEach(id =>
    document.getElementById(id).classList.add('hidden')
  );

  // Validate: at least one location field must be filled (skip for Fundable)
  const selCountries = countriesMS.getSelected();
  const selStates = statesMS.getSelected();
  const selCities = citiesMS.getSelected();
  const hasAny = selCountries.includes('Any');
  const leadSource = document.querySelector('input[name="lead_source"]:checked')?.value || 'supabase';
  if (leadSource !== 'fundable' && !hasAny && selCountries.length === 0 && selStates.length === 0 && selCities.length === 0) {
    showError('Missing location', 'Select at least one Country, State, or City to filter results.');
    runBtn.disabled = false;
    cancelBtn.classList.add('hidden');
    loading.classList.add('hidden');
    document.dispatchEvent(new CustomEvent('runComplete'));
    return;
  }

  const sizes = sizesMS.getSelected();

  const payload = {
    lead_source: leadSource,
    employee_count: sizes,
    country: selCountries,
    state: selStates,
    city: selCities,
    industry: industryMS.getSelected(),
    sub_industry: subIndustryMS.getSelected(),
    role: selectedRoles.length > 0 ? selectedRoles : null,
    max_leads: parseInt(document.getElementById('max_leads').value) || 5,
    request_description: document.getElementById('request_description').value.trim() || null,
    product_description: document.getElementById('product_description').value.trim() || null,
    intent_signals: document.getElementById('intent_signals').value.trim() || null,
    exclude_companies: excludedCompanies,
  };

  // Fundable-specific fields
  if (leadSource === 'fundable') {
    const fd = document.getElementById('funding_days').value;
    if (fd) payload.funding_days = parseInt(fd);
    const ft = document.getElementById('financing_types').value.trim();
    if (ft) payload.financing_types = ft;
  }

  // Poll progress during discovery (shows Fundable company-by-company progress)
  let discoverPollId = null;
  if (leadSource === 'fundable') {
    discoverPollId = setInterval(async () => {
      try {
        const pr = await fetch('/api/progress');
        const pg = await pr.json();
        if (pg.step) statusMsg.textContent = pg.step;
      } catch(e) {}
    }, 1500);
  }

  try {
    statusMsg.textContent = 'Sending request...';
    const resp = await fetch('/api/discover-phases', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify(payload),
      signal: discoverAbort.signal,
    });
    if (discoverPollId) clearInterval(discoverPollId);

    if (!resp.ok) {
      const errText = await resp.text().catch(() => '');
      let errMsg = 'Server returned ' + resp.status;
      try { const j = JSON.parse(errText); if (j.error) errMsg = j.error; } catch(e) {}
      showError('Discovery failed', errMsg);
      return;
    }

    const data = await resp.json();

    if (data.error) {
      showError('Discovery failed', data.error);
      return;
    }

    // Phase 0: Auto-select industries in dropdown
    const p0 = document.getElementById('phase0');
    p0.classList.remove('hidden');
    document.getElementById('phase0Count').textContent =
      data.extracted_industries.length + ' industries, ' + data.extracted_sub_industries.length + ' sub-industries';
    const p0Body = document.getElementById('phase0Body');
    p0Body.innerHTML = '<p style="font-size:13px;color:#6b7d93;margin-bottom:8px;">Industries auto-selected in dropdown above. Sub-industries expanded automatically.</p>';

    // Auto-select industries and sub-industries in dropdowns
    industryMS.setSelected(data.extracted_industries);
    // Expand subs for extracted industries
    const autoSubs = new Set();
    data.extracted_industries.forEach(ind => {
      (INDUSTRY_SUBS[ind] || []).forEach(s => autoSubs.add(s));
    });
    subIndustryMS.setOptions([...autoSubs].sort());
    subIndustryMS.setSelected(data.extracted_sub_industries);

    // Auto-populate sizes and locations from LLM extraction
    if ((data.extracted_sizes || []).length && sizesMS.getSelected().length === 0) {
      sizesMS.setSelected(data.extracted_sizes);
    }
    const curCountries = countriesMS.getSelected();
    console.log('[DEBUG] curCountries:', curCountries, 'extracted:', data.extracted_countries);
    if ((data.extracted_countries || []).length && (curCountries.length === 0 || (curCountries.length === 1 && curCountries[0] === 'Any'))) {
      console.log('[DEBUG] Setting countries:', data.extracted_countries);
      countriesMS.setSelected(data.extracted_countries);
      onCountryChange(data.extracted_countries);
    } else {
      console.log('[DEBUG] Skipped country update');
    }
    if ((data.extracted_states || []).length && statesMS.getSelected().length === 0) {
      statesMS.setSelected(data.extracted_states);
    }
    if ((data.extracted_cities || []).length && citiesMS.getSelected().length === 0) {
      citiesMS.setSelected(data.extracted_cities);
    }

    // Phase 1: Unique roles from DB
    statusMsg.textContent = 'Fetching roles from DB...';
    const p1 = document.getElementById('phase1');
    p1.classList.remove('hidden');
    document.getElementById('phase1Count').textContent = data.unique_roles_total + ' unique roles';
    document.getElementById('phase1Info').textContent =
      'Found ' + data.unique_roles_total + ' unique roles in DB matching industry + location + size filters.' +
      (data.message ? ' ' + data.message : '');

    if (data.unique_roles_total === 0) {
      showWarning('No roles found in database', 'No leads matched your industry + location + size filters. Try broadening your filters.');
      statusMsg.textContent = 'No results';
      return;
    }

    // Phase 2: LLM scored roles
    statusMsg.textContent = 'Scoring roles...';
    const p2 = document.getElementById('phase2');
    p2.classList.remove('hidden');
    document.getElementById('phase2Count').textContent = data.ranked_roles.length + ' matched roles';
    const roleList = document.getElementById('roleList');
    roleList.innerHTML = '';
    data.ranked_roles.forEach((role, i) => {
      const score = data.scored_roles[role] || 0;
      const chip = document.createElement('div');
      chip.className = 'role-chip';
      chip.innerHTML = '<span class="rank">#' + (i+1) + '</span>' + role + ' <span style="color:#5a9fd4;font-size:11px;margin-left:4px">' + score.toFixed(1) + '</span>';
      roleList.appendChild(chip);
    });

    if (data.ranked_roles.length === 0) {
      showWarning('No matching roles', data.message || 'LLM could not match any roles to your description. Try adjusting the buyer profile or manually adding roles.');
      statusMsg.textContent = 'No matching roles';
      return;
    }

    // Phase 3
    statusMsg.textContent = 'Done';
    const p3 = document.getElementById('phase3');
    p3.classList.remove('hidden');
    document.getElementById('phase3Count').textContent =
      data.companies.length + ' companies (target: ' + (payload.max_leads * 5) + ')';
    const tbody = document.getElementById('companiesBody');
    tbody.innerHTML = '';
    data.companies.forEach((c, i) => {
      const tr = document.createElement('tr');
      const webUrl = c.website ? (c.website.startsWith('http') ? c.website : 'https://'+c.website) : '';
      const liUrl = c.company_linkedin ? (c.company_linkedin.startsWith('http') ? c.company_linkedin : 'https://'+c.company_linkedin) : '';
      tr.innerHTML =
        '<td style="color:#4a5c73">' + (i+1) + '</td>' +
        '<td class="company-name">' + (c.company_name || '-') + '</td>' +
        '<td>' + (c.location || '-') + '</td>' +
        '<td>' + (webUrl ? '<a href="' + webUrl + '" target="_blank" class="company-link">' + c.website + '</a>' : '-') + '</td>' +
        '<td>' + (liUrl ? '<a href="' + liUrl + '" target="_blank" class="company-link">LinkedIn</a>' : '-') + '</td>' +
        '<td>' + (c.employee_count || '-') + '</td>' +
        '<td>' + (c.industry || '-') + (c.sub_industry ? '<br><span style="color:#5a6b80;font-size:11px">' + c.sub_industry + '</span>' : '') + '</td>' +
        '<td style="max-width:400px;font-size:12px;color:#6b7d93;">' + (c.description || '-') + '</td>' +
        '<td style="font-size:12px;color:#c5d0dc;">' + (c.unique_roles || []).join(', ') + '</td>' +
        '<td class="fit-score">' + (c.fit_score || 0).toFixed(2) + '</td>' +
        '<td><span class="leads-count">' + c.lead_count + ' leads</span></td>';
      tbody.appendChild(tr);
    });

    // Store companies for intent check
    window._phase3Companies = data.companies;
    window._rankedRoles = data.ranked_roles;
    window._roleScores = data.scored_roles || {};

    // Full Run: auto-trigger Verify + Intent Check
    if (document.getElementById('fullRunCheck').checked && data.companies.length > 0) {
      window._fullRunPending = true;
      statusMsg.textContent = 'Full Run — starting Verify + Intent Check...';
      loading.classList.remove('hidden');
      runBtn.disabled = true;
      cancelBtn.classList.add('hidden');
      setTimeout(() => runProcessCompanies(), 100);
      return;
    }

  } catch (err) {
    if (err.name !== 'AbortError') {
      showError('Request failed', err.message === 'Failed to fetch' ? 'Could not connect to server. Make sure it\\'s running on localhost:5002.' : err.message);
    }
  } finally {
    runBtn.disabled = false;
    cancelBtn.classList.add('hidden');
    loading.classList.add('hidden');
    if (!window._fullRunPending) {
      document.dispatchEvent(new CustomEvent('runComplete'));
    }
    window._fullRunPending = false;
  }
});

function proceedAnyway() {
  document.getElementById('processBtn').scrollIntoView({behavior: 'smooth'});
}

// ── Expand Size Range ──
function expandAndRetry() {
  const currentSizes = sizesMS.getSelected();
  const allSizes = ALL_SIZES;

  // Find min and max index of current selection
  let minIdx = allSizes.length;
  let maxIdx = -1;
  currentSizes.forEach(s => {
    const idx = allSizes.indexOf(s);
    if (idx >= 0) {
      minIdx = Math.min(minIdx, idx);
      maxIdx = Math.max(maxIdx, idx);
    }
  });

  // Expand ±1
  const newSizes = [...currentSizes];
  if (minIdx > 0 && !newSizes.includes(allSizes[minIdx - 1])) {
    newSizes.push(allSizes[minIdx - 1]);
  }
  if (maxIdx < allSizes.length - 1 && !newSizes.includes(allSizes[maxIdx + 1])) {
    newSizes.push(allSizes[maxIdx + 1]);
  }

  // Update dropdown
  sizesMS.setSelected(newSizes);
  showWarning('Size range expanded', 'Now includes: ' + newSizes.join(', ') + '. Click "Verify Leads" to retry.');
}

// ── Process Companies (Verify + Intent per company) ──

document.getElementById('processCancelBtn').addEventListener('click', () => {
  if (window._processAbort) { window._processAbort.abort(); window._processAbort = null; }
  showWarning('Cancelled', 'Verify + Intent Check was cancelled.');
  document.getElementById('processStatusMsg').textContent = 'Cancelled';
  document.getElementById('processBtn').disabled = false;
  document.getElementById('processCancelBtn').classList.add('hidden');
  document.getElementById('processLoading').classList.add('hidden');
});

async function runProcessCompanies() {
  const companies = window._phase3Companies;
  if (!companies || companies.length === 0) {
    showError('No companies to verify', 'Run Discovery first to find companies.');
    document.dispatchEvent(new CustomEvent('runComplete'));
    return;
  }

  const btn = document.getElementById('processBtn');
  const loadEl = document.getElementById('processLoading');
  const statusEl = document.getElementById('processStatusMsg');
  const processCancelBtn = document.getElementById('processCancelBtn');
  btn.disabled = true;
  processCancelBtn.classList.remove('hidden');
  loadEl.classList.remove('hidden');
  statusEl.textContent = 'Starting...';
  document.getElementById('phase4').classList.add('hidden');
  document.getElementById('phase5').classList.add('hidden');

  if (window._processAbort) { window._processAbort.abort(); }
  window._processAbort = new AbortController();

  // Poll progress every 2 seconds — also detect completion
  let progressDone = false;
  const progressInterval = setInterval(async () => {
    try {
      const pr = await fetch('/api/progress');
      const pg = await pr.json();
      if (pg.status === 'running') {
        statusEl.textContent =
          'Company ' + pg.companies_done + '/' + pg.companies_total +
          ': ' + (pg.company || '...') +
          ' | ' + (pg.step || '') +
          ' | Verified: ' + pg.prospects + '/' + pg.target +
          ' | Leads: ' + pg.leads +
          ' | Cost: $' + (pg.cost_usd || 0).toFixed(3);
      } else if (pg.status === 'done' && !progressDone) {
        progressDone = true;
        statusEl.textContent = 'Done! Leads: ' + pg.leads + ' | Prospects: ' + pg.prospects;
        // Render results from progress data if main fetch hasn't returned
        if (pg.leads_pool || pg.prospect_pool) {
          renderResults({
            leads_pool: pg.leads_pool || [],
            prospect_pool: pg.prospect_pool || [],
            intent_results: pg.intent_results || [],
            stats: pg.stats || {},
          });
        }
      }
    } catch(e) {}
  }, 2000);

  const sizes = sizesMS.getSelected();

  const payload = {
    lead_source: document.querySelector('input[name="lead_source"]:checked')?.value || 'supabase',
    companies: companies,
    industries: industryMS.getSelected(),
    ranked_roles: window._rankedRoles || [],
    countries: countriesMS.getSelected().filter(c => c !== 'Any'),
    states: statesMS.getSelected(),
    cities: citiesMS.getSelected(),
    employee_counts: sizes,
    max_leads: parseInt(document.getElementById('max_leads').value) || 5,
    product_description: document.getElementById('product_description').value.trim() || '',
    request_description: document.getElementById('request_description').value.trim() || '',
    intent_signals: document.getElementById('intent_signals').value.trim() || '',
    icp_filters: {
      industries: industryMS.getSelected(),
      countries: countriesMS.getSelected().filter(c => c !== 'Any'),
      states: statesMS.getSelected(),
      cities: citiesMS.getSelected(),
      employee_counts: sizes,
    },
    role_scores: window._roleScores || {},
    skip_stage4: document.getElementById('skipStage4').checked,
    exclude_companies: excludedCompanies,
  };

  try {
    const resp = await fetch('/api/process-companies', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify(payload),
      signal: window._processAbort.signal,
    });
    const data = await resp.json();

    if (data.error) {
      showError('Processing failed', data.error);
      return;
    }

    const stats = data.stats;

    // Phase 4: Show intent results per company
    renderResults(data);

    // Auto-save results to server when Full Run is active
    if (document.getElementById('fullRunCheck').checked) {
      setTimeout(() => {
        saveResults().catch(e => console.error('Auto-save failed:', e));
      }, 500);
    }

  } catch (err) {
    if (err.name !== 'AbortError') {
      showError('Processing failed', err.message);
    }
  } finally {
    clearInterval(progressInterval);
    btn.disabled = false;
    processCancelBtn.classList.add('hidden');
    loadEl.classList.add('hidden');
    document.dispatchEvent(new CustomEvent('runComplete'));
  }
}

function renderResults(data) {
    const stats = data.stats || {};

    // Show funding results if available
    if (data.funding_results && data.funding_results.length > 0) {
      const fp = document.getElementById('fundingPhase');
      fp.classList.remove('hidden');
      const funded = data.funding_results.filter(f => f.funded);
      document.getElementById('fundingCount').textContent =
        funded.length + ' funded / ' + (data.funding_results.length - funded.length) + ' not funded';
      const ftbody = document.getElementById('fundingBody');
      ftbody.innerHTML = '';
      data.funding_results.forEach((f, i) => {
        const tr = document.createElement('tr');
        tr.style.background = f.funded ? '#0f1f15' : '';
        tr.innerHTML =
          '<td style="color:#4a5c73">' + (i+1) + '</td>' +
          '<td style="color:#e2e8f0;font-weight:600;">' + (f.company_name || '-') + '</td>' +
          '<td style="color:' + (f.funded ? '#4c4' : '#c44') + ';font-weight:600;">' + (f.funded ? 'FUNDED' : 'NOT FUNDED') + '</td>' +
          '<td style="font-size:12px;color:#a8c0dc;">' + (f.evidence || '-') + '</td>';
        ftbody.appendChild(tr);
      });
    }

    window._intentResults = data.intent_results || [];
    if (data.intent_results && data.intent_results.length > 0) {
      const p4 = document.getElementById('phase4');
      p4.classList.remove('hidden');
      document.getElementById('phase4Count').textContent =
        stats.companies_intent_checked + ' companies | ' +
        stats.leads_count + '/' + stats.target_leads + ' leads | ' +
        'Email: ' + stats.email_passed + ' | Stage4: ' + stats.stage4_passed;

      const tbody = document.getElementById('intentBody');
      tbody.innerHTML = '';
      data.intent_results.forEach((c, i) => {
        const tr = document.createElement('tr');
        tr.innerHTML =
          '<td style="color:#4a5c73">' + (i+1) + '</td>' +
          '<td class="company-name">' + (c.company_name || '-') + '</td>' +
          '<td>-</td><td>-</td><td>-</td><td>-</td><td>-</td><td>-</td><td>-</td><td>-</td>' +
          '<td style="font-size:12px;color:#c5d0dc;max-width:500px;">' + (c.intent_paragraph || '-') + '</td>' +
          '<td class="fit-score">' + (c.intent_score || 0).toFixed(2) + '</td>' +
          '<td class="fit-score">' + (c.lead_score || 0).toFixed(2) + '</td>';
        tbody.appendChild(tr);
      });
    }

    // Phase 5: Two pools
    const leadsPool = data.leads_pool || [];
    const prospectPool = data.prospect_pool || [];
    window._finalLeads = leadsPool;
    window._prospectLeads = prospectPool;

    if (leadsPool.length > 0 || prospectPool.length > 0) {
      const p5 = document.getElementById('phase5');
      p5.classList.remove('hidden');
      document.getElementById('phase5Count').textContent =
        leadsPool.length + ' leads + ' + prospectPool.length + ' prospects';

      function renderLeadRows(tbodyId, leads) {
        const tbody = document.getElementById(tbodyId);
        tbody.innerHTML = '';
        const hasFunding = leads.some(l => l.funding_type || l.funding_date);
        // Add funding headers if needed
        const thead = tbody.closest('table').querySelector('thead tr');
        if (hasFunding && !thead.querySelector('.funding-col')) {
          const cols = ['Funding Type', 'Funding Date', 'Days Since', 'Total Raised', 'Investors'];
          cols.forEach(c => { const th = document.createElement('th'); th.className = 'funding-col'; th.textContent = c; thead.insertBefore(th, thead.querySelector('th:nth-last-child(3)')); });
        }
        leads.forEach((l, i) => {
          const tr = document.createElement('tr');
          let html =
            '<td style="color:#4a5c73">' + (i+1) + '</td>' +
            '<td class="company-name">' + (l.name || '-') + '</td>' +
            '<td>' + (l.email || '-') + '</td>' +
            '<td>' + (l.role || '-') + '</td>' +
            '<td>' + (l.company || '-') + '</td>' +
            '<td>' + (l.linkedin ? '<a href="' + l.linkedin + '" target="_blank" class="company-link">LI</a>' : '-') + '</td>' +
            '<td>' + (l.website ? '<a href="' + (l.website.startsWith('http') ? l.website : 'https://'+l.website) + '" target="_blank" class="company-link">Web</a>' : '-') + '</td>' +
            '<td>' + (l.company_linkedin ? '<a href="' + (l.company_linkedin.startsWith('http') ? l.company_linkedin : 'https://'+l.company_linkedin) + '" target="_blank" class="company-link">Co LI</a>' : '-') + '</td>' +
            '<td>' + (l.phone || '-') + '</td>' +
            '<td>' + (l.industry || '-') + '</td>' +
            '<td>' + (l.sub_industry || '-') + '</td>' +
            '<td>' + (l.city || '-') + '</td>' +
            '<td>' + (l.state || '-') + '</td>' +
            '<td>' + (l.country || '-') + '</td>' +
            '<td>' + (l.hq_city || '-') + '</td>' +
            '<td>' + (l.hq_state || '-') + '</td>' +
            '<td>' + (l.hq_country || '-') + '</td>' +
            '<td>' + (l.employee_count || '-') + '</td>' +
            '<td style="font-size:12px;color:#6b7d93;max-width:300px;">' + (l.description || '-') + '</td>';
          if (hasFunding) {
            html +=
              '<td>' + (l.funding_type || '-') + '</td>' +
              '<td>' + (l.funding_date || '-') + '</td>' +
              '<td>' + (l.days_since_funding || '-') + '</td>' +
              '<td>' + (l.total_raised ? '$' + Number(l.total_raised).toLocaleString() : '-') + '</td>' +
              '<td style="font-size:12px;max-width:200px;">' + (l.funding_investors || '-') + '</td>';
          }
          html +=
            '<td style="font-size:12px;color:#c5d0dc;max-width:400px;">' + (l.intent_details || '-') + '</td>' +
            '<td class="fit-score">' + (l.lead_score_pct || '0%') + '</td>' +
            '<td style="font-size:11px;color:#7a8ba0;max-width:500px;">' + (l.evidence || '-').replace(/[|]/g, '<br>') + '</td>';
          tr.innerHTML = html;
          tbody.appendChild(tr);
        });
      }

      renderLeadRows('leadsPoolBody', leadsPool);
      renderLeadRows('prospectPoolBody', prospectPool);
    }
}

// ── Download All Leads from Phase 3 ──
async function downloadAllLeadsExcel() {
  const companies = window._phase3Companies || [];
  if (!companies.length) {
    showWarning('Nothing to export', 'No companies available to export.');
    return;
  }

  // Flatten all leads from all companies
  const allLeads = [];
  companies.forEach(c => {
    (c.leads || []).forEach(l => {
      allLeads.push({
        name: ((l.first_name || '') + ' ' + (l.last_name || '')).trim(),
        email: l.email || '',
        role: l.role || '',
        company: l.company_name || '',
        linkedin: l.linkedin || '',
        website: l.website || '',
        company_linkedin: l.company_linkedin || '',
        phone: l.phone || '',
        industry: l.industry || '',
        sub_industry: l.sub_industry || '',
        city: l.city || '',
        state: l.state || '',
        country: l.country || '',
        hq_city: l.hq_city || '',
        hq_state: l.hq_state || '',
        hq_country: l.hq_country || '',
        employee_count: l.employee_count || '',
        description: l.description || '',
      });
    });
  });

  if (!allLeads.length) {
    showWarning('Nothing to export', 'No leads attached to companies.');
    return;
  }

  try {
    const resp = await fetch('/api/download-leads', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({leads: allLeads, prospects: []}),
    });
    if (!resp.ok) {
      showError('Download failed', 'Server returned ' + resp.status);
      return;
    }
    const blob = await resp.blob();
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url;
    a.download = 'all_leads_raw.xlsx';
    document.body.appendChild(a);
    a.click();
    document.body.removeChild(a);
    URL.revokeObjectURL(url);
  } catch (err) {
    showError('Download failed', err.message);
  }
}

// ── Download Table as CSV ──
function downloadTableCSV(phaseId) {
  let data, filename, headers;

  if (phaseId === 'phase3') {
    data = window._phase3Companies || [];
    filename = 'companies.csv';
    headers = ['Company', 'Location', 'Website', 'LinkedIn', 'Size', 'Industry', 'Sub Industry', 'Description', 'Roles', 'Fit Score', 'Lead Count'];
    const rows = data.map(c => [
      c.company_name, c.location, c.website, c.company_linkedin, c.employee_count,
      c.industry, c.sub_industry, c.description, (c.unique_roles||[]).join('; '),
      c.fit_score, c.lead_count
    ]);
    exportCSV(headers, rows, filename);
  } else if (phaseId === 'phase4') {
    data = window._intentResults || [];
    filename = 'intent_results.csv';
    headers = ['Company', 'Intent Paragraph', 'Intent Score', 'Lead Score'];
    const rows = data.map(c => [
      c.company_name, c.intent_paragraph, c.intent_score, c.lead_score
    ]);
    exportCSV(headers, rows, filename);
  }
}

function exportCSV(headers, rows, filename) {
  const escape = (v) => '"' + String(v || '').replace(/"/g, '""') + '"';
  let csv = headers.map(escape).join(',') + '\\n';
  rows.forEach(row => {
    csv += row.map(escape).join(',') + '\\n';
  });
  const blob = new Blob([csv], {type: 'text/csv'});
  const url = URL.createObjectURL(blob);
  const a = document.createElement('a');
  a.href = url;
  a.download = filename;
  document.body.appendChild(a);
  a.click();
  document.body.removeChild(a);
  URL.revokeObjectURL(url);
}

// ── Save Results to Server ──
async function saveResults() {
  const leads = window._finalLeads || [];
  const prospects = window._prospectLeads || [];
  if (leads.length === 0 && prospects.length === 0) return;

  const clientName = document.getElementById('client_name').value.trim();
  const buyer = document.getElementById('request_description').value.trim();
  const product = document.getElementById('product_description').value.trim();
  const label = clientName || buyer || product || 'export';

  try {
    const resp = await fetch('/api/save-results', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({leads: leads, prospects: prospects, label: label}),
    });
    const result = await resp.json();
    if (result.saved) {
      showWarning('Results saved', 'File saved to results/' + result.saved);
    } else {
      console.error('Save failed:', result.error);
    }
  } catch (err) {
    console.error('Save results failed:', err);
  }
}

// ── Download Excel ──
async function downloadExcel() {
  const leads = window._finalLeads || [];
  const prospects = window._prospectLeads || [];
  if (leads.length === 0 && prospects.length === 0) {
    showWarning('Nothing to download', 'No leads available to export.');
    return;
  }

  try {
    const resp = await fetch('/api/download-leads', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({leads: leads, prospects: prospects}),
    });

    if (!resp.ok) {
      const err = await resp.json().catch(() => ({}));
      showError('Download failed', err.error || 'Server returned ' + resp.status);
      return;
    }

    const blob = await resp.blob();
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url;
    const ts = new Date().toISOString().slice(0,16).replace('T','_').replace(':','-');
    a.download = 'leads_export_' + ts + '.xlsx';
    document.body.appendChild(a);
    a.click();
    document.body.removeChild(a);
    URL.revokeObjectURL(url);
  } catch (err) {
    showError('Download failed', err.message);
  }
}

// ── Clear All Fields ──
function clearAllFields() {
  sizesMS.setSelected([]);
  countriesMS.setSelected([]);
  statesMS.setOptions([]); statesMS.setSelected([]);
  citiesMS.setOptions([]); citiesMS.setSelected([]);
  industryMS.setSelected([]);
  subIndustryMS.setOptions([]); subIndustryMS.setSelected([]);
  selectedRoles = [];
  renderRoleChips();
  excludedCompanies = [];
  renderExcludeChips();
  document.getElementById('client_name').value = '';
  document.getElementById('request_description').value = '';
  document.getElementById('product_description').value = '';
  document.getElementById('intent_signals').value = '';
  document.getElementById('max_leads').value = '5';
  ['phase0','phase1','phase2','phase3','phase4','phase5'].forEach(id => {
    document.getElementById(id).classList.add('hidden');
  });
}

// ── Paste Brief Modal ──
function openPasteModal() {
  document.getElementById('pasteBriefText').value = '';
  document.getElementById('pasteModal').classList.add('open');
  setTimeout(() => document.getElementById('pasteBriefText').focus(), 50);
}

function closePasteModal() {
  document.getElementById('pasteModal').classList.remove('open');
}

document.getElementById('pasteModal').addEventListener('mousedown', (e) => {
  if (e.target.classList.contains('modal-overlay')) closePasteModal();
});

function parsePastedBrief() {
  const raw = document.getElementById('pasteBriefText').value.trim();
  if (!raw) { closePasteModal(); return; }

  try {

  const lines = raw.split('\\n').map(l => l.trim());

  const sections = {};
  let currentKey = null;
  let currentLines = [];

  const sectionPatterns = [
    { key: 'buyer', patterns: ['buyer profile', 'lead company description', 'target profile', 'buyer'] },
    { key: 'product', patterns: ['product/service', 'product description', 'product', 'service'] },
    { key: 'intent', patterns: ['intent signals', 'intent', 'signals'] },
    { key: 'countries', patterns: ['countries', 'country', 'location'] },
    { key: 'sizes', patterns: ['company sizes', 'company size', 'sizes', 'size', 'employees'] },
    { key: 'roles', patterns: ['roles', 'role', 'target roles'] },
    { key: 'states', patterns: ['states', 'state'] },
    { key: 'cities', patterns: ['cities', 'city'] },
  ];

  function detectSection(line) {
    const colonIdx = line.indexOf(':');
    const label = colonIdx > 0 ? line.substring(0, colonIdx).trim().toLowerCase() : line.replace(/[:;]/g, '').trim().toLowerCase();
    for (const sp of sectionPatterns) {
      if (sp.patterns.some(p => label === p)) return { key: sp.key, value: colonIdx > 0 ? line.substring(colonIdx + 1).trim() : '' };
    }
    return null;
  }

  for (const line of lines) {
    if (!line) continue;
    const detected = detectSection(line);
    if (detected) {
      if (currentKey) sections[currentKey] = currentLines.join(' ').trim();
      currentKey = detected.key;
      currentLines = [];
      if (detected.value) currentLines.push(detected.value);
    } else if (currentKey) {
      currentLines.push(line);
    }
  }
  if (currentKey) sections[currentKey] = currentLines.join(' ').trim();

  if (sections.buyer) {
    document.getElementById('request_description').value = sections.buyer;
  }
  if (sections.product) {
    document.getElementById('product_description').value = sections.product;
  }
  if (sections.intent) {
    document.getElementById('intent_signals').value = sections.intent;
  }

  if (sections.countries) {
    const countryText = sections.countries.toLowerCase();
    const matched = [];
    for (const c of ALL_COUNTRIES) {
      if (countryText.includes(c.toLowerCase())) matched.push(c);
    }
    if (matched.length) {
      countriesMS.setSelected(matched);
      onCountryChange(matched);
    }
  }

  if (sections.states) {
    const stateList = sections.states.split(/[,;]+/).map(s => s.trim()).filter(Boolean);
    const matched = [];
    for (const s of stateList) {
      const found = ALL_STATES.find(st => st.toLowerCase() === s.toLowerCase());
      if (found) matched.push(found);
    }
    if (matched.length) {
      statesMS.setSelected(matched);
      onStateChange(matched);
    }
  }

  if (sections.cities) {
    updateCityOptions();
    const cityList = sections.cities.split(/[,;]+/).map(c => c.trim()).filter(Boolean);
    const available = citiesMS.allOptions;
    const matched = [];
    for (const c of cityList) {
      const found = available.find(av => av.toLowerCase() === c.toLowerCase());
      if (found) matched.push(found);
    }
    if (matched.length) citiesMS.setSelected(matched);
  }

  if (sections.sizes) {
    const sizeText = sections.sizes.replace(/[\u2013\u2014]/g, '-').replace(/employees/gi, '').replace(/,/g, '');
    const inputRanges = sizeText.match(/\\d+[\\s]*-[\\s]*\\d+|\\d+\\+/g) || [];
    const matched = [];
    for (const input of inputRanges) {
      const clean = input.replace(/\\s/g, '');
      const inputParts = clean.match(/(\\d+)-(\\d+)/);
      if (!inputParts) continue;
      const inLo = parseInt(inputParts[1]);
      const inHi = parseInt(inputParts[2]);
      for (const s of ALL_SIZES) {
        if (matched.includes(s)) continue;
        const norm = s.replace(/,/g, '');
        const sParts = norm.match(/(\\d+)-(\\d+)/);
        if (!sParts) continue;
        const sLo = parseInt(sParts[1]);
        const sHi = parseInt(sParts[2]);
        if (Math.abs(inLo - sLo) <= 1 && Math.abs(inHi - sHi) <= 1) {
          matched.push(s);
          break;
        }
      }
    }
    if (matched.length) sizesMS.setSelected(matched);
  }

  if (sections.roles) {
    const roleList = sections.roles.split(/[,;]+/).map(r => r.trim()).filter(Boolean);
    selectedRoles = roleList;
    renderRoleChips();
  }

  saveFormState();
  } catch(e) { console.error('Paste brief parsing error:', e); }
  closePasteModal();
}
</script>
<script>
(function() {
  let serverId = null;
  async function checkReload() {
    try {
      const r = await fetch('/api/livereload');
      const d = await r.json();
      if (serverId === null) { serverId = d.id; return; }
      if (d.id !== serverId) location.reload();
    } catch(e) {}
  }
  setInterval(checkReload, 1500);
})();
</script>
</body>
</html>
"""


@app.route("/")
def index():
    return render_template_string(HTML)


@app.route("/api/parse-icp", methods=["POST"])
def parse_icp_endpoint():
    """Parse ICP text inputs into structured fields."""
    from target_fit_model.icp_extract import parse_icp
    data = request.get_json()
    result = parse_icp(
        icp_description=data.get("icp_description", ""),
        product_description=data.get("product_description", ""),
        intent_signals=data.get("intent_signals", ""),
    )
    return jsonify(result)


def _discover_fundable(data):
    """Fundable discovery path. Replaces Phases 1-3 with Fundable API."""
    from fundable_query import discover_fundable_leads
    from target_fit_model.scoring import compute_fit_score

    industries = data.get("industry", [])
    sub_industries = data.get("sub_industry", [])
    raw_sizes = data.get("employee_count", [])
    raw_countries = data.get("country", [])
    countries = [c for c in raw_countries if c != "Any"]
    states = data.get("state", [])
    cities = data.get("city", [])
    max_leads = data.get("max_leads", 5)
    request_desc = data.get("request_description")
    product_desc = data.get("product_description")
    intent_sigs = data.get("intent_signals")
    funding_days = data.get("funding_days")
    financing_types = data.get("financing_types", "")

    # Phase 0: ICP parsing if no industries selected
    if not industries and request_desc:
        _progress.update({"status": "running", "step": "Extracting ICP via Claude...", "leads": 0, "prospects": 0})
        from target_fit_model.icp_extract import parse_icp as extract_icp_filters
        icp_result = extract_icp_filters(request_desc, product_desc, intent_sigs)
        if icp_result:
            logger.info(f"[Fundable ICP] Parsed: industries={icp_result.get('industries')}, "
                        f"employee_counts={icp_result.get('employee_counts')}, "
                        f"countries={icp_result.get('countries')}, "
                        f"states={icp_result.get('states')}, cities={icp_result.get('cities')}")
            industries = icp_result.get("industries", [])
            sub_industries = icp_result.get("sub_industries", [])
            if not raw_sizes and icp_result.get("employee_counts"):
                raw_sizes = icp_result["employee_counts"]
            if not countries and icp_result.get("countries"):
                countries = icp_result["countries"]
            if not states and icp_result.get("states"):
                states = icp_result["states"]
            if not cities and icp_result.get("cities"):
                cities = icp_result["cities"]

    if not industries:
        return jsonify({"error": "No industries found. Select manually or provide a Buyer Profile."}), 400

    # Fundable discovery
    target_companies = max_leads * COMPANY_MULTIPLIER
    _progress.update({"status": "running", "step": "Searching Fundable API...", "leads": 0, "prospects": 0, "target": max_leads})

    def _fundable_progress(msg):
        _progress["step"] = msg

    companies = discover_fundable_leads(
        industries=industries,
        employee_counts=raw_sizes,
        countries=countries,
        states=states,
        cities=cities,
        funding_days=funding_days,
        financing_types=financing_types,
        max_companies=target_companies,
        progress_callback=_fundable_progress,
    )

    exclude_companies = {c.strip().lower() for c in data.get("exclude_companies", []) if c.strip()}
    if exclude_companies:
        companies = [c for c in companies if (c.get("company_name") or "").strip().lower() not in exclude_companies]

    # Build role_scores (static 1.0 for all Fundable roles)
    all_roles = set()
    for c in companies:
        for l in c.get("leads", []):
            r = (l.get("role") or "").strip()
            if r:
                all_roles.add(r)
    role_scores = {r: 1.0 for r in all_roles}

    # Compute fit_score for each lead
    icp = {
        "industries": industries,
        "sub_industries": sub_industries,
        "roles": list(all_roles),
        "countries": countries,
        "states": states,
        "cities": cities,
        "employee_counts": raw_sizes,
    }
    for company in companies:
        for lead in company.get("leads", []):
            fit_score, breakdown = compute_fit_score(lead, icp)
            lead["fit_score"] = fit_score
            lead["fit_breakdown"] = breakdown
        if company["leads"]:
            company["fit_score"] = round(max(l["fit_score"] for l in company["leads"]), 4)

    companies.sort(key=lambda c: c.get("fit_score", 0), reverse=True)

    _progress.update({"status": "done", "step": f"Found {len(companies)} companies"})

    return jsonify({
        "error": None,
        "lead_source": "fundable",
        "extracted_industries": industries,
        "extracted_sub_industries": sub_industries,
        "extracted_sizes": raw_sizes,
        "extracted_countries": countries,
        "extracted_states": states,
        "extracted_cities": cities,
        "unique_roles_total": len(all_roles),
        "scored_roles": role_scores,
        "ranked_roles": sorted(all_roles),
        "companies": companies,
    })


@app.route("/api/discover-phases", methods=["POST"])
def discover_phases():
    """Run Phase 0-3: extract industries, query DB, rank roles, collect companies."""
    try:
        return _discover_phases_inner()
    except Exception as e:
        logger.error(f"[Discovery] Unhandled error: {e}", exc_info=True)
        return jsonify({"error": f"Server error: {str(e)}"}), 500


def _discover_phases_inner():
    data = request.get_json()

    # Branch to Fundable path if selected
    lead_source = data.get("lead_source", "supabase")
    if lead_source == "fundable":
        return _discover_fundable(data)

    # Expand employee count to include DB variants (with/without "employees" suffix)
    raw_sizes = data.get("employee_count", [])
    employee_counts = []
    for s in raw_sizes:
        employee_counts.extend(_EMPLOYEE_VARIANTS.get(s, [s]))

    raw_countries = data.get("country", [])
    # "Any" means no country filter
    countries = [c for c in raw_countries if c != "Any"]
    states = data.get("state", [])
    cities = data.get("city", [])
    industry_input = data.get("industry", [])
    sub_industry_input = data.get("sub_industry", [])
    max_leads = data.get("max_leads", 5)
    request_desc = data.get("request_description")
    product_desc = data.get("product_description")
    intent_sigs = data.get("intent_signals")
    role_input = data.get("role")
    exclude_companies = data.get("exclude_companies", [])

    if not request_desc and not industry_input:
        return jsonify({"error": "Either select an Industry or provide a Lead Company Description"}), 400

    # Detect funding intent → fetch 100x companies instead of 5x
    from funding_check import detect_funding_intent
    has_funding = detect_funding_intent(intent_sigs or "")
    if has_funding:
        target_companies = max_leads * 100
        logger.info(f"[Discovery] Funding intent detected — fetching {target_companies} companies (100x)")
    else:
        target_companies = max_leads * COMPANY_MULTIPLIER

    # ── Phase 0: Determine industries ──
    # If user selected industries manually → use those, skip LLM
    # If no industries selected → extract from description via LLM
    if industry_input:
        industries = industry_input
        sub_industries = sub_industry_input if sub_industry_input else []
        # If no subs selected, expand from mapping
        if not sub_industries:
            for ind in industries:
                sub_industries.extend(_INDUSTRY_SUBS.get(ind, []))
    else:
        icp_filters = extract_icp_filters(
            icp_description=request_desc,
            product_description=product_desc,
            intent_signals=intent_sigs,
        )
        industries = icp_filters.get("industries", [])
        sub_industries = icp_filters.get("sub_industries", [])

        # Auto-fill employee counts and locations from LLM if user didn't select manually
        if not employee_counts:
            llm_sizes = icp_filters.get("employee_counts", [])
            for s in llm_sizes:
                employee_counts.extend(_EMPLOYEE_VARIANTS.get(s, [s]))
        if not countries:
            countries = icp_filters.get("countries", [])
        if not states:
            states = icp_filters.get("states", [])
        if not cities:
            cities = icp_filters.get("cities", [])

    if not industries:
        # Retry once
        logger.warning("[Discovery] No industries found, retrying...")
        icp_filters = extract_icp_filters(request_desc, product_desc, intent_sigs)
        industries = icp_filters.get("industries", [])
        sub_industries = icp_filters.get("sub_industries", [])

    if not industries:
        return jsonify({
            "error": "Could not extract industries from description. Please be more specific.",
            "extracted_industries": [], "extracted_sub_industries": [],
        })

    # Compute display-friendly sizes for frontend auto-populate
    reverse_variants = {}
    for display, variants in _EMPLOYEE_VARIANTS.items():
        for v in variants: reverse_variants[v] = display
    extracted_sizes = list(dict.fromkeys(reverse_variants.get(ec, ec) for ec in employee_counts))

    # ── Phase 1: Fetch unique roles from DB ──
    unique_roles = fetch_unique_roles(
        industries=industries,
        sub_industries=sub_industries,
        countries=countries,
        states=states,
        cities=cities,
        employee_counts=employee_counts,
    )

    if not unique_roles:
        return jsonify({
            "error": None,
            "extracted_industries": industries,
            "extracted_sub_industries": sub_industries[:20],
            "extracted_sizes": extracted_sizes,
            "extracted_countries": countries,
            "extracted_states": states,
            "extracted_cities": cities,
            "unique_roles_total": 0,
            "scored_roles": {},
            "ranked_roles": [],
            "companies": [],
        })

    # ── Phase 2: Role matching ──
    if isinstance(role_input, list):
        target_roles = [r.strip() for r in role_input if r.strip()]
    elif isinstance(role_input, str):
        target_roles = [r.strip() for r in role_input.split(",") if r.strip()]
    else:
        target_roles = []

    if target_roles:
        # User provided roles — LLM matches same/similar roles only
        role_scores = score_roles_via_llm(
            target_roles=target_roles,
            candidate_roles=unique_roles,
            request_description=request_desc,
            product_description=product_desc,
            intent_signals=intent_sigs,
        )
        ranked_roles = get_ranked_roles(role_scores, target_roles)
        logger.info(f"[Phase2] LLM matched {len(ranked_roles)} similar roles from {len(unique_roles)}")
    else:
        # No roles provided — use LLM to discover from description/intent
        role_scores = score_roles_via_llm(
            target_roles=[],
            candidate_roles=unique_roles,
            request_description=request_desc,
            product_description=product_desc,
            intent_signals=intent_sigs,
        )
        ranked_roles = get_ranked_roles(role_scores)

    if not ranked_roles:
        return jsonify({
            "error": None,
            "extracted_industries": industries,
            "extracted_sub_industries": sub_industries[:20],
            "extracted_sizes": extracted_sizes,
            "extracted_countries": countries,
            "extracted_states": states,
            "extracted_cities": cities,
            "unique_roles_total": len(unique_roles),
            "scored_roles": {},
            "ranked_roles": [],
            "companies": [],
            "message": "No matching roles found. Try different roles or broaden filters.",
        })

    # ── Phase 3: Fetch leads batch by batch until 5x unique companies ──
    icp = {
        "industries": industries,
        "sub_industries": sub_industries,
        "roles": ranked_roles,
        "countries": countries,
        "states": states,
        "cities": cities,
        "employee_counts": employee_counts,
    }

    seen_companies = {c.strip().lower() for c in exclude_companies if c.strip()}
    seen_lead_ids = set()
    companies = []
    all_scored_leads = []
    BATCH_SIZE = 10  # Fetch 10 roles at a time
    role_idx = 0

    while len(companies) < target_companies and role_idx < len(ranked_roles):
        batch_roles = ranked_roles[role_idx:role_idx + BATCH_SIZE]
        role_idx += BATCH_SIZE

        batch_leads = fetch_leads_for_roles(
            industries=industries,
            sub_industries=sub_industries,
            role_strings=batch_roles,
            countries=countries,
            states=states,
            cities=cities,
            employee_counts=employee_counts,
            limit=target_companies * 5,
        )

        # Dedupe against previously fetched leads
        new_leads = []
        for lead in batch_leads:
            lid = lead.get("lead_id")
            if lid and lid not in seen_lead_ids:
                seen_lead_ids.add(lid)
                new_leads.append(lead)

        # Score new leads
        for lead in new_leads:
            fit_score, breakdown = compute_fit_score(lead, icp)
            all_scored_leads.append({**lead, "fit_score": fit_score, "fit_breakdown": breakdown})

        # Collect unique companies from new leads
        for lead in new_leads:
            company_key = (lead.get("company_name") or "").strip().lower()
            if not company_key or company_key in seen_companies:
                continue

            seen_companies.add(company_key)

            # Count all leads for this company across all batches
            company_leads = [
                l for l in all_scored_leads
                if (l.get("company_name") or "").strip().lower() == company_key
            ]

            # Build location string from HQ fields
            hq_parts = []
            hq_city = (lead.get("hq_city") or "").strip()
            hq_state = (lead.get("hq_state") or "").strip()
            hq_country = (lead.get("hq_country") or "").strip()
            if hq_city:
                hq_parts.append(hq_city)
            if hq_state:
                hq_parts.append(hq_state)
            if hq_country:
                hq_parts.append(hq_country)
            hq_location = ", ".join(hq_parts) if hq_parts else ""

            # Fallback to lead location if no HQ
            if not hq_location:
                loc_parts = []
                if (lead.get("city") or "").strip():
                    loc_parts.append(lead["city"].strip())
                if (lead.get("state") or "").strip():
                    loc_parts.append(lead["state"].strip())
                if (lead.get("country") or "").strip():
                    loc_parts.append(lead["country"].strip())
                hq_location = ", ".join(loc_parts)

            # Collect unique roles for this company
            company_roles = sorted(set(
                (l.get("role") or "").strip()
                for l in company_leads
                if (l.get("role") or "").strip()
            ))

            # Include raw leads for this company (for verify + intent step)
            raw_leads_for_company = []
            for cl in company_leads:
                # Extract phone from lead_blob before discarding it
                blob = cl.get("lead_blob") or {}
                if isinstance(blob, str):
                    try:
                        import json as _j
                        blob = _j.loads(blob)
                    except Exception:
                        blob = {}
                phone_numbers = blob.get("phone_numbers", [])
                phone = ", ".join(phone_numbers) if phone_numbers else ""

                raw_leads_for_company.append({
                    "lead_id": cl.get("lead_id"),
                    "first_name": cl.get("first_name"),
                    "last_name": cl.get("last_name"),
                    "email": cl.get("email"),
                    "role": cl.get("role"),
                    "company_name": cl.get("company_name"),
                    "linkedin": cl.get("linkedin"),
                    "website": cl.get("website"),
                    "company_linkedin": cl.get("company_linkedin"),
                    "industry": cl.get("industry"),
                    "sub_industry": cl.get("sub_industry"),
                    "city": cl.get("city"),
                    "state": cl.get("state"),
                    "country": cl.get("country"),
                    "hq_city": cl.get("hq_city"),
                    "hq_state": cl.get("hq_state"),
                    "hq_country": cl.get("hq_country"),
                    "employee_count": cl.get("employee_count"),
                    "description": cl.get("description"),
                    "phone": phone,
                })

            logger.info(f"[Phase3] Company '{lead.get('company_name')}': {len(raw_leads_for_company)} raw leads attached")

            companies.append({
                "company_name": lead.get("company_name"),
                "location": hq_location,
                "website": lead.get("website"),
                "company_linkedin": lead.get("company_linkedin"),
                "industry": lead.get("industry"),
                "sub_industry": lead.get("sub_industry"),
                "employee_count": lead.get("employee_count"),
                "description": lead.get("description") or "",
                "unique_roles": company_roles,
                "fit_score": round(company_leads[0]["fit_score"], 4) if company_leads else 0,
                "lead_count": len(company_leads),
                "top_role": lead.get("role"),
                "role_score": role_scores.get(lead.get("role"), 0),
                "leads": raw_leads_for_company,
            })

            if len(companies) >= target_companies:
                break

    # Sort companies by fit_score descending — best fit gets verified first
    companies.sort(key=lambda c: c.get("fit_score", 0), reverse=True)

    # Format scored roles for display
    scored_roles_display = {r: round(s, 2) for r, s in sorted(role_scores.items(), key=lambda x: -x[1])}

    # Map back employee_counts to display format for UI auto-population
    extracted_sizes = []
    reverse_variants = {}
    for display, variants in _EMPLOYEE_VARIANTS.items():
        for v in variants:
            reverse_variants[v] = display
    for ec in employee_counts:
        display = reverse_variants.get(ec, ec)
        if display not in extracted_sizes:
            extracted_sizes.append(display)

    return jsonify({
        "error": None,
        "extracted_industries": industries,
        "extracted_sub_industries": sub_industries[:20],
        "extracted_sizes": extracted_sizes,
        "extracted_countries": countries,
        "extracted_states": states,
        "extracted_cities": cities,
        "unique_roles_total": len(unique_roles),
        "scored_roles": scored_roles_display,
        "ranked_roles": ranked_roles[:50],
        "companies": companies,
    })


# Global progress state for polling
_progress = {"status": "idle", "detail": "", "company": "", "step": "", "leads": 0, "prospects": 0, "target": 0, "companies_done": 0, "companies_total": 0, "leads_pool": [], "prospect_pool": [], "intent_results": [], "stats": {}}


@app.route("/api/progress", methods=["GET"])
def get_progress():
    from target_fit_model.openrouter import get_cost_tracker
    cost = get_cost_tracker()
    return jsonify({**_progress, "cost_usd": round(cost["total_usd"], 4)})


@app.route("/api/process-companies", methods=["POST"])
def process_companies():
    """
    Combined verify + intent: process one company at a time.
    Updates global _progress dict for polling.
    """
    try:
        return _process_companies_inner()
    except Exception as e:
        logger.error(f"[Process] Unhandled error: {e}", exc_info=True)
        return jsonify({"error": f"Server error: {str(e)}"}), 500


def _process_companies_inner():
    import asyncio
    from truelist_verify import verify_emails_batch, verify_emails_inline
    from stage4_person_verification import run_lead_validation_stage4
    from target_fit_model.intent_enrichment import research_company_intent, compute_lead_score, _cross_source_boost
    from evidence import build_evidence, compose_intent_details

    global _progress
    data = request.get_json()
    companies = data.get("companies", [])
    industries = data.get("industries", [])
    ranked_roles = data.get("ranked_roles", [])
    countries = data.get("countries", [])
    states = data.get("states", [])
    cities = data.get("cities", [])
    employee_counts = data.get("employee_counts", [])
    max_leads = data.get("max_leads", 5)
    product_desc = data.get("product_description", "")
    request_desc = data.get("request_description", "")
    intent_sigs = data.get("intent_signals", "")
    icp_filters = data.get("icp_filters", {})
    role_scores = data.get("role_scores", {})

    lead_source = data.get("lead_source", "supabase")

    exclude_companies = {c.strip().lower() for c in data.get("exclude_companies", []) if c.strip()}
    if exclude_companies:
        companies = [c for c in companies if (c.get("company_name") or "").strip().lower() not in exclude_companies]

    if not companies:
        return jsonify({"error": "No companies to process"}), 400

    from target_fit_model.config import INTENT_THRESHOLD, GEMINI_MODEL, COMPANY_MULTIPLIER, ICP_PARSER_MODEL, PERPLEXITY_MODEL, PERPLEXITY_TIMEOUT
    from target_fit_model.openrouter import chat_completion, chat_completion_json, reset_cost_tracker, get_cost_tracker
    skip_stage4 = data.get("skip_stage4", False)
    reset_cost_tracker()

    # ── Funding filter: if intent mentions funding, filter companies first ──
    # Skip for Fundable — funding data is already verified
    from funding_check import detect_funding_intent, classify_funding_criteria, check_company_funding, _parse_date_to_sortkey
    has_funding_intent = detect_funding_intent(intent_sigs) if lead_source != "fundable" else False
    funding_results = []
    if has_funding_intent:
        _progress.update({"status": "running", "step": "Funding intent detected! Classifying criteria...", "leads": 0, "prospects": 0, "target": max_leads, "companies_done": 0, "companies_total": len(companies)})
        funding_criteria = classify_funding_criteria(intent_sigs)
        target_funded = max_leads * 2
        funded_companies = []
        checked_company_names = set()
        total_checked = 0

        for company in companies:
            if len(funded_companies) >= target_funded:
                break
            cn = company.get("company_name", "")
            if not cn or cn.lower() in checked_company_names:
                continue
            checked_company_names.add(cn.lower())
            total_checked += 1
            _progress["step"] = f"Funding check: {cn} ({total_checked}) — {len(funded_companies)}/{target_funded} funded"
            _progress["companies_done"] = total_checked

            passed, evidence, date_str = check_company_funding(cn, company.get("website", ""), funding_criteria)
            funding_results.append({"company_name": cn, "funded": passed, "evidence": evidence})

            if passed:
                company["_funding_passed"] = True
                company["_funding_evidence"] = evidence
                company["_funding_date"] = date_str
                funded_companies.append(company)

        if funded_companies:
            funded_companies.sort(key=lambda c: _parse_date_to_sortkey(c.get("_funding_date", "")), reverse=True)
            companies = funded_companies
            logger.info(f"[FundingCheck] Found {len(funded_companies)} funded from {total_checked} checked")
        else:
            logger.warning("[FundingCheck] No funded companies found")

    # Broaden intent signals via LLM before processing any company
    original_signal_list = [s.strip() for s in intent_sigs.split(",") if s.strip()] if intent_sigs else []
    expanded_intent = intent_sigs
    if intent_sigs and product_desc:
        _progress.update({"status": "running", "step": "Expanding intent signals via LLM...", "leads": 0, "prospects": 0, "target": max_leads, "companies_done": 0, "companies_total": len(companies)})
        try:
            expand_prompt = f"""You are a B2B sales intelligence expert. A company sells this product:

Product: {product_desc}

The user wants to find companies showing these buying signals:
{intent_sigs}

Expand this list significantly. Think about ALL possible signals that could indicate a company needs this product. Include:
- Direct signals (exactly what was listed)
- Indirect signals (related behaviors that suggest need)
- Growth signals (expansion, funding, new markets, headcount growth)
- Leadership signals (new executives, restructuring, leadership changes)
- Digital signals (new tools adopted, website changes, job postings)
- Content signals (blog posts, social media, webinars about related topics)
- Financial signals (revenue changes, funding rounds, acquisitions)
- Competitive signals (competitors using similar products)
- Pain signals (negative reviews, complaints, challenges mentioned publicly)

IMPORTANT: Keep the user's original signals FIRST and unchanged. Then add 10-15 high-quality expanded signals. Total 20-25 signals. Quality over quantity — only add signals that are realistically detectable via public web search. No explanations, just the comma-separated list."""

            expanded = chat_completion(
                prompt=expand_prompt,
                model=ICP_PARSER_MODEL,
                temperature=0.3,
                max_tokens=2000,
            )
            if expanded and len(expanded) > len(intent_sigs):
                expanded_intent = expanded.strip()
                logger.info(f"[Process] Expanded intent signals: {len(intent_sigs)} chars → {len(expanded_intent)} chars")
        except Exception as e:
            logger.error(f"[Process] Intent expansion failed: {e}")

    leads_pool = []
    prospect_pool = []
    intent_results = []

    _progress.update({"status": "running", "leads": 0, "prospects": 0, "target": max_leads, "companies_done": 0, "companies_total": len(companies)})
    stats = {
        "companies_processed": 0,
        "companies_with_verified_leads": 0,
        "companies_intent_checked": 0,
        "total_leads_fetched": 0,
        "email_passed": 0,
        "stage4_passed": 0,
        "intent_failed": 0,
    }

    async def verify_one_lead(lead):
        try:
            lead_dict = {
                "full_name": f"{(lead.get('first_name') or '').strip()} {(lead.get('last_name') or '').strip()}".strip(),
                "company": lead.get("company_name", ""),
                "linkedin": lead.get("linkedin", ""),
                "city": lead.get("city", ""),
                "state": lead.get("state", ""),
                "country": lead.get("country", ""),
                "role": lead.get("role", ""),
                "email": lead.get("email", ""),
            }
            result = await run_lead_validation_stage4(lead_dict)
            return lead, result
        except Exception as e:
            logger.error(f"[Process] Stage 4 failed for {lead.get('email')}: {e}")
            return lead, {"passed": True, "rejection_reason": None, "data": {}}

    target_verified_companies = max_leads * COMPANY_MULTIPLIER
    verified_company_count = 0

    # ── ICP Description Check: batch validate companies against ICP ──
    ICP_CHECK_BATCH = 15
    icp_target = max_leads * 20  # need 20x target leads worth of valid companies
    icp_summary = f"Buyer Profile: {request_desc}"

    _progress.update({"step": f"ICP check: validating {len(companies)} companies...", "leads": 0})
    logger.info(f"[Process] ICP description check: {len(companies)} companies, target {icp_target} valid")

    icp_checked = []
    for icp_i in range(0, len(companies), ICP_CHECK_BATCH):
        if len(icp_checked) >= icp_target:
            logger.info(f"[Process] ICP check: reached {len(icp_checked)} valid companies, stopping")
            break

        icp_batch = companies[icp_i:icp_i+ICP_CHECK_BATCH]
        icp_items = "\n".join(
            f"{j+1}. {c.get('company_name','?')} — {(c.get('description') or 'no description')[:100]}"
            for j, c in enumerate(icp_batch)
        )
        icp_batch_num = icp_i // ICP_CHECK_BATCH + 1
        _progress["step"] = f"ICP check batch {icp_batch_num}: {len(icp_checked)} valid so far..."

        icp_result = None
        for attempt in range(2):
            try:
                icp_result = chat_completion_json(
                    prompt=f"Which of these companies match this ICP? A company matches if its business is the type described in the ICP.\n\nICP:\n{icp_summary}\n\nCOMPANIES:\n{icp_items}\n\nFor each return: company_name, match (true/false), reason (1 sentence). JSON array only.",
                    model=GEMINI_MODEL, temperature=0, max_tokens=3000,
                )
                if icp_result and isinstance(icp_result, list):
                    break
            except Exception as e:
                logger.error(f"[Process] ICP batch {icp_batch_num} attempt {attempt+1} failed: {e}")

        if icp_result and isinstance(icp_result, list):
            matched = {(r.get("company_name") or "").strip().lower() for r in icp_result if r.get("match")}
            for c in icp_batch:
                if (c.get("company_name") or "").strip().lower() in matched:
                    icp_checked.append(c)
            logger.info(f"[Process] ICP batch {icp_batch_num}: {len(matched)}/{len(icp_batch)} matched")
        else:
            logger.error(f"[Process] ICP batch {icp_batch_num} failed after 2 attempts, passing batch through")
            icp_checked.extend(icp_batch)

    logger.info(f"[Process] ICP check done: {len(icp_checked)}/{len(companies)} companies passed")
    _progress["step"] = f"ICP check: {len(icp_checked)} companies match ICP"
    companies = icp_checked

    # ── Chunked pre-filter + deep check: 25 at a time until target met ──
    CHUNK_SIZE = 25
    all_companies = list(companies)
    chunk_offset = 0
    processed_names = set()
    use_prefilter = bool(intent_sigs and expanded_intent)

    logger.info(f"[Process] Starting chunked processing: {len(all_companies)} companies, target {max_leads} leads")

    while len(leads_pool) < max_leads and chunk_offset < len(all_companies):
        chunk = all_companies[chunk_offset:chunk_offset + CHUNK_SIZE]
        chunk_offset += CHUNK_SIZE
        round_num = chunk_offset // CHUNK_SIZE

        if not chunk:
            break

        # Pre-filter this chunk
        if use_prefilter:
            _progress["step"] = f"Round {round_num}: pre-filtering {len(chunk)} companies..."
            pf_items = "\n".join(f"{j+1}. {c.get('company_name','?')} ({c.get('website','')})" for j, c in enumerate(chunk))
            pf_names = set()
            try:
                pf_result = chat_completion_json(
                    prompt=f"Which of these companies show ANY of these buying signals in the last 6 months?\n\nSIGNALS: {expanded_intent[:500]}\n\nCOMPANIES:\n{pf_items}\n\nFor each match return: company_name, signal, reason (1 sentence). Only include companies with real recent evidence. JSON array only.",
                    model=PERPLEXITY_MODEL, system_prompt="Search for real evidence. JSON array only.",
                    temperature=0, max_tokens=4000, timeout=PERPLEXITY_TIMEOUT)
                if pf_result and isinstance(pf_result, list):
                    for r in pf_result:
                        n = (r.get("company_name") or "").strip().lower()
                        if n: pf_names.add(n)
                    logger.info(f"[Process] Round {round_num}: pre-filter {len(pf_names)}/{len(chunk)} matched")
                else:
                    # No results — keep all from this chunk
                    for c in chunk: pf_names.add((c.get("company_name") or "").strip().lower())
            except Exception as e:
                logger.error(f"[Process] Round {round_num} pre-filter failed: {e}")
                for c in chunk: pf_names.add((c.get("company_name") or "").strip().lower())

            chunk = [c for c in chunk if (c.get("company_name") or "").strip().lower() in pf_names]

        _progress["step"] = f"Round {round_num}: deep checking {len(chunk)} companies ({len(leads_pool)}/{max_leads} leads)"
        logger.info(f"[Process] Round {round_num}: deep checking {len(chunk)} companies")

        # Deep check each company in this chunk
        for company in chunk:
            if len(leads_pool) >= max_leads:
                logger.info(f"[Process] Target reached: {len(leads_pool)} leads")
                break

            cn = (company.get("company_name") or "").strip().lower()
            if cn in processed_names:
                continue
            processed_names.add(cn)

            company_name = (company.get("company_name") or "").strip()
            if not company_name:
                continue

            stats["companies_processed"] += 1
            _progress.update({"company": company_name, "step": "Processing leads", "companies_done": stats["companies_processed"], "leads": len(leads_pool), "prospects": verified_company_count, "target": target_verified_companies})
            logger.info(f"[Process] Company {stats['companies_processed']}/{len(companies)}: {company_name} "
                         f"({len(leads_pool)}/{max_leads} leads)")

            # Step 1: Use leads already fetched in Phase 3 (no re-query)
            company_leads = company.get("leads", [])

            stats["total_leads_fetched"] += len(company_leads)
            if not company_leads:
                continue

            # Step 2: Email verification (skip for Fundable — already verified)
            if lead_source == "fundable":
                email_passed = [l for l in company_leads if l.get("email")]
                for lead in email_passed:
                    lead["_email_status"] = "email_ok"
                stats["email_passed"] += len(email_passed)
            else:
                _progress["step"] = f"Verifying emails ({len(company_leads)} leads)"
                emails = [l.get("email") for l in company_leads if l.get("email")]
                email_results = {}
                # Verify one by one (inline API)
                for email in emails:
                    try:
                        result = verify_emails_inline([email])
                        email_results.update(result)
                    except Exception as e:
                        logger.error(f"[Process] Truelist failed for {email}: {e}")

                email_passed = []
                for lead in company_leads:
                    email = lead.get("email", "")
                    if not email:
                        continue
                    result = email_results.get(email, {})
                    if result.get("passed", False):
                        lead["_email_status"] = result.get("status", "unknown")
                        email_passed.append(lead)
                stats["email_passed"] += len(email_passed)

            # Email domain check: email domain must match company domain
            company_domain = (company.get("website") or "").lower().replace("https://","").replace("http://","").replace("www.","").rstrip("/").split("/")[0]
            if company_domain:
                domain_checked = []
                for lead in email_passed:
                    email = lead.get("email", "")
                    if not email:
                        continue
                    email_domain = email.split("@")[1] if "@" in email else ""
                    if not email_domain or email_domain == company_domain or email_domain.endswith("." + company_domain) or company_domain.endswith("." + email_domain):
                        domain_checked.append(lead)
                    else:
                        logger.info(f"[Process] Email domain mismatch: {email} vs {company_domain}")
                email_passed = domain_checked

            if not email_passed:
                continue

            # Step 3: Stage 4 person verification (skip for Fundable — already verified)
            if lead_source == "fundable" or skip_stage4:
                verified = email_passed
                for lead in verified:
                    lead["_stage4_passed"] = True
                stats["stage4_passed"] += len(verified)
            else:
                _progress["step"] = f"Stage 4 verification ({len(email_passed)} leads)"
                verified = []
                try:
                    loop = asyncio.new_event_loop()
                    asyncio.set_event_loop(loop)
                    s4_results = loop.run_until_complete(
                        asyncio.gather(*[verify_one_lead(l) for l in email_passed])
                    )
                    loop.close()
                except Exception as e:
                    logger.error(f"[Process] Stage 4 batch failed for {company_name}: {e}")
                    s4_results = [(l, {"passed": True}) for l in email_passed]

                for lead, s4_result in s4_results:
                    stats["stage4_passed"] += 1 if s4_result.get("passed") else 0
                    if s4_result.get("passed", False):
                        lead["_stage4_passed"] = True
                        lead["_stage4_data"] = s4_result.get("data", {})
                        verified.append(lead)

            if not verified:
                continue

            stats["companies_with_verified_leads"] += 1
            verified_company_count += 1

            # Step 4: Perplexity intent check for this company (only if has verified leads)
            _progress["step"] = f"Intent check ({len(verified)} verified leads)"
            intent_data = None
            if intent_sigs:
                stats["companies_intent_checked"] += 1
                try:
                    intent_data = research_company_intent(
                        company=company,
                        product_description=product_desc,
                        request_description=request_desc,
                        intent_signals=expanded_intent,
                    )
                except Exception as e:
                    logger.error(f"[Process] Intent failed for {company_name}: {e}")
                    stats["intent_failed"] += 1

            # Compute scores
            if intent_data and isinstance(intent_data, dict):
                signals = intent_data.get("signals", [])
                # Recompute intent score using our tiered hybrid formula
                from target_fit_model.intent_enrichment import compute_intent_score_from_signals
                raw_intent = compute_intent_score_from_signals(signals, original_signal_list)
                fit_score = float(intent_data.get("fit_score", 0.5))
                boosted_intent, source_count = _cross_source_boost(signals, raw_intent)
                lead_score = compute_lead_score(boosted_intent, fit_score)
                intent_paragraph = intent_data.get("intent_paragraph", "")
                data_gaps = intent_data.get("_data_gaps", [])
            else:
                boosted_intent = 0.5
                fit_score = 0.5
                source_count = 0
                lead_score = 0.5
                intent_paragraph = ""
                data_gaps = []
                signals = []

            company_intent = {
                "intent_score": boosted_intent,
                "fit_score_perplexity": fit_score,
                "lead_score": lead_score,
                "intent_paragraph": intent_paragraph,
                "intent_signals_detail": signals,
                "_source_count": source_count,
                "_data_gaps": data_gaps,
                "company_name": company_name,
            }
            intent_results.append(company_intent)

            company_description = (company.get("description") or "").strip()

            # Keep only the best role-fit lead per company
            verified.sort(key=lambda l: role_scores.get((l.get("role") or "").strip(), 0), reverse=True)
            verified = verified[:1]

            # Build final leads for this company
            for lead in verified:

                first = (lead.get("first_name") or "").strip()
                last = (lead.get("last_name") or "").strip()
                name = f"{first} {last}".strip()

                # Phone already extracted in Phase 3
                phone = lead.get("phone", "")

                # HQ location
                hq_parts = [
                    (lead.get("hq_city") or "").strip(),
                    (lead.get("hq_state") or "").strip(),
                    (lead.get("hq_country") or "").strip(),
                ]
                hq_location = ", ".join(p for p in hq_parts if p)

                # Build evidence breakdown
                from evidence import build_evidence
                lead_dict_for_evidence = {
                    "country": lead.get("country"),
                    "state": lead.get("state"),
                    "city": lead.get("city"),
                    "industry": lead.get("industry"),
                    "sub_industry": lead.get("sub_industry"),
                    "employee_count": lead.get("employee_count"),
                    "role": lead.get("role"),
                }
                evidence = build_evidence(
                    lead=lead_dict_for_evidence,
                    icp_filters=icp_filters,
                    role_scores=role_scores,
                    intent_data=company_intent,
                    verification_data={
                        "email_status": lead.get("_email_status", "unknown"),
                        "stage4_passed": True,
                    },
                    data_gaps=data_gaps,
                )

                # Strip URLs from intent_paragraph → clean Intent Details
                # Extract URLs → append to Evidence as Sources section
                import re as _re
                _url_pattern = r'https?://[^\s\)\]\},;\"\'<>]+'
                source_urls = list(dict.fromkeys(_re.findall(_url_pattern, intent_paragraph)))
                clean_intent = _re.sub(r'\s*\(?' + _url_pattern + r'\)?\s*', ' ', intent_paragraph).strip()
                clean_intent = _re.sub(r'\s{2,}', ' ', clean_intent)  # collapse double spaces
                clean_intent = _re.sub(r'\s+([.,;])', r'\1', clean_intent)  # fix space before punctuation

                if source_urls:
                    evidence += "\n---\nSources: " + ", ".join(source_urls)

                lead_entry = {
                    "name": name,
                    "email": lead.get("email"),
                    "role": lead.get("role"),
                    "company": company_name,
                    "linkedin": lead.get("linkedin"),
                    "website": lead.get("website"),
                    "company_linkedin": lead.get("company_linkedin"),
                    "phone": phone,
                    "industry": lead.get("industry"),
                    "sub_industry": lead.get("sub_industry"),
                    "city": lead.get("city"),
                    "state": lead.get("state"),
                    "country": lead.get("country"),
                    "hq_city": lead.get("hq_city"),
                    "hq_state": lead.get("hq_state"),
                    "hq_country": lead.get("hq_country"),
                    "employee_count": lead.get("employee_count"),
                    "description": company_description,
                    "intent_details": clean_intent,
                    "lead_score": lead_score,
                    "lead_score_pct": f"{lead_score:.0%}",
                    "evidence": evidence,
                    "intent_score": boosted_intent,
                }

                # Fundable funding columns
                if lead_source == "fundable":
                    lead_entry["funding_type"] = company.get("_funding_type", "")
                    lead_entry["funding_date"] = company.get("_funding_date", "")
                    lead_entry["funding_amount"] = company.get("_funding_amount", "")
                    lead_entry["funding_investors"] = company.get("_funding_investors", "")
                    lead_entry["total_raised"] = company.get("_total_raised", "")
                    lead_entry["days_since_funding"] = company.get("_days_since_funding", "")

                # Split by intent threshold
                if boosted_intent >= INTENT_THRESHOLD:
                    leads_pool.append(lead_entry)
                    logger.info(f"[Process] → LEAD ({len(leads_pool)}/{max_leads}): {name} at {company_name} (intent={boosted_intent:.2f})")
                else:
                    prospect_pool.append(lead_entry)
                    logger.info(f"[Process] → PROSPECT: {name} at {company_name} (intent={boosted_intent:.2f} < {INTENT_THRESHOLD})")

    final_cost = round(get_cost_tracker()["total_usd"], 4)
    # Sort by lead_score descending
    leads_pool.sort(key=lambda l: l.get("lead_score", 0), reverse=True)
    prospect_pool.sort(key=lambda l: l.get("lead_score", 0), reverse=True)

    final_stats = {
        **stats,
        "leads_count": len(leads_pool),
        "prospect_count": len(prospect_pool),
        "target_leads": max_leads,
        "enough": len(leads_pool) >= max_leads,
        "intent_threshold": INTENT_THRESHOLD,
        "cost_usd": final_cost,
        "cost_breakdown": get_cost_tracker()["by_model"],
    }

    _progress.update({
        "status": "done", "step": "Complete",
        "leads": len(leads_pool), "prospects": len(prospect_pool),
        "leads_pool": leads_pool, "prospect_pool": prospect_pool,
        "intent_results": intent_results,
        "stats": final_stats,
    })

    return jsonify({
        "error": None,
        "leads_pool": leads_pool,
        "prospect_pool": prospect_pool,
        "intent_results": intent_results,
        "funding_results": funding_results if has_funding_intent else None,
        "stats": final_stats,
    })


@app.route("/api/download-leads", methods=["POST"])
def download_leads():
    """Download leads as Excel file."""
    from flask import send_file
    import io

    try:
        import openpyxl
    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    data = request.get_json()
    leads = data.get("leads", [])
    prospects = data.get("prospects", [])

    if not leads and not prospects:
        return jsonify({"error": "No leads to download"}), 400

    wb = openpyxl.Workbook()

    headers = [
        "Name", "Email", "Role", "Company", "LinkedIn", "Website",
        "Company LinkedIn", "Phone", "Industry", "Sub Industry",
        "Contact City", "Contact State", "Contact Country",
        "Company HQ City", "Company HQ State", "Company HQ Country",
        "Employee Count", "Description", "Intent Details", "Lead Score", "Evidence",
    ]

    # Add funding columns if any lead has funding data
    all_leads = leads + prospects
    has_funding = any(l.get("funding_type") or l.get("funding_date") for l in all_leads)
    if has_funding:
        # Insert funding columns before "Intent Details"
        idx = headers.index("Intent Details")
        for col in reversed(["Funding Type", "Funding Date", "Days Since Funding", "Total Raised", "Investors"]):
            headers.insert(idx, col)

    for lead in leads:
        lead["_type"] = "Lead"
    for lead in prospects:
        lead["_type"] = "Prospect"
    combined = leads + prospects

    headers.insert(0, "Type")
    ws = wb.active
    ws.title = "All Leads"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for lead in combined:
        row = [lead.get("_type", "")]
        row.extend([
            lead.get("name", ""), lead.get("email", ""), lead.get("role", ""),
            lead.get("company", ""), lead.get("linkedin", ""), lead.get("website", ""),
            lead.get("company_linkedin", ""), lead.get("phone", ""),
            lead.get("industry", ""), lead.get("sub_industry", ""),
            lead.get("city", lead.get("contact_city", "")),
            lead.get("state", lead.get("contact_state", "")),
            lead.get("country", lead.get("contact_country", "")),
            lead.get("hq_city", ""), lead.get("hq_state", ""), lead.get("hq_country", ""),
            lead.get("employee_count", ""), lead.get("description", ""),
        ])
        if has_funding:
            row.extend([
                lead.get("funding_type", ""), lead.get("funding_date", ""),
                lead.get("days_since_funding", ""), lead.get("total_raised", ""),
                lead.get("funding_investors", ""),
            ])
        row.extend([
            lead.get("intent_details", ""),
            lead.get("lead_score_pct", lead.get("lead_score", "")),
            lead.get("evidence", ""),
        ])
        ws.append(row)
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="leads_export.xlsx",
    )


@app.route("/api/save-results", methods=["POST"])
def save_results():
    """Save leads to an Excel file in the results/ directory."""
    import re

    try:
        import openpyxl
    except ImportError:
        return jsonify({"error": "openpyxl not installed"}), 500

    data = request.get_json()
    leads = data.get("leads", [])
    prospects = data.get("prospects", [])
    label = data.get("label", "")

    if not leads and not prospects:
        return jsonify({"error": "No leads to save"}), 400

    results_dir = os.path.join(os.path.dirname(__file__), "results")
    os.makedirs(results_dir, exist_ok=True)

    slug = re.sub(r'[^a-zA-Z0-9]+', '_', label.strip())[:80].strip('_').lower() or "export"
    base = slug
    path = os.path.join(results_dir, f"{base}.xlsx")
    counter = 2
    while os.path.exists(path):
        path = os.path.join(results_dir, f"{base}_{counter}.xlsx")
        counter += 1

    wb = openpyxl.Workbook()
    headers = [
        "Name", "Email", "Role", "Company", "LinkedIn", "Website",
        "Company LinkedIn", "Phone", "Industry", "Sub Industry",
        "Contact City", "Contact State", "Contact Country",
        "Company HQ City", "Company HQ State", "Company HQ Country",
        "Employee Count", "Description", "Intent Details", "Lead Score", "Evidence",
    ]

    all_leads = leads + prospects
    has_funding = any(l.get("funding_type") or l.get("funding_date") for l in all_leads)
    if has_funding:
        idx = headers.index("Intent Details")
        for col in reversed(["Funding Type", "Funding Date", "Days Since Funding", "Total Raised", "Investors"]):
            headers.insert(idx, col)

    def write_sheet(ws, sheet_leads):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        for lead in sheet_leads:
            row = [
                lead.get("name", ""), lead.get("email", ""), lead.get("role", ""),
                lead.get("company", ""), lead.get("linkedin", ""), lead.get("website", ""),
                lead.get("company_linkedin", ""), lead.get("phone", ""),
                lead.get("industry", ""), lead.get("sub_industry", ""),
                lead.get("city", lead.get("contact_city", "")),
                lead.get("state", lead.get("contact_state", "")),
                lead.get("country", lead.get("contact_country", "")),
                lead.get("hq_city", ""), lead.get("hq_state", ""), lead.get("hq_country", ""),
                lead.get("employee_count", ""), lead.get("description", ""),
            ]
            if has_funding:
                row.extend([
                    lead.get("funding_type", ""), lead.get("funding_date", ""),
                    lead.get("days_since_funding", ""), lead.get("total_raised", ""),
                    lead.get("funding_investors", ""),
                ])
            row.extend([
                lead.get("intent_details", ""),
                lead.get("lead_score_pct", lead.get("lead_score", "")),
                lead.get("evidence", ""),
            ])
            ws.append(row)
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    for lead in leads:
        lead["_type"] = "Lead"
    for lead in prospects:
        lead["_type"] = "Prospect"
    combined = leads + prospects

    headers.insert(0, "Type")
    ws = wb.active
    ws.title = "All Leads"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for lead in combined:
        row = [lead.get("_type", "")]
        row.extend([
            lead.get("name", ""), lead.get("email", ""), lead.get("role", ""),
            lead.get("company", ""), lead.get("linkedin", ""), lead.get("website", ""),
            lead.get("company_linkedin", ""), lead.get("phone", ""),
            lead.get("industry", ""), lead.get("sub_industry", ""),
            lead.get("city", lead.get("contact_city", "")),
            lead.get("state", lead.get("contact_state", "")),
            lead.get("country", lead.get("contact_country", "")),
            lead.get("hq_city", ""), lead.get("hq_state", ""), lead.get("hq_country", ""),
            lead.get("employee_count", ""), lead.get("description", ""),
        ])
        if has_funding:
            row.extend([
                lead.get("funding_type", ""), lead.get("funding_date", ""),
                lead.get("days_since_funding", ""), lead.get("total_raised", ""),
                lead.get("funding_investors", ""),
            ])
        row.extend([
            lead.get("intent_details", ""),
            lead.get("lead_score_pct", lead.get("lead_score", "")),
            lead.get("evidence", ""),
        ])
        ws.append(row)
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(path)
    filename = os.path.basename(path)
    return jsonify({"saved": filename, "path": path})


_server_id = str(id(app)) + str(os.getpid())


@app.route("/api/livereload")
def livereload():
    return jsonify({"id": _server_id})


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5002, debug=True)
